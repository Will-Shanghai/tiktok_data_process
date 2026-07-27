# -*- coding: utf-8 -*-
"""
TikTok 越南周报自动生成脚本
- 从飞书表格读取越南店 SKU 映射及成本配置
- 仅处理越南站点
- 当前支持越南本土店、越南跨境店
- 使用固定越南汇率，不读取其他国家配置
- 输出四张 Excel 报表（周汇总、分类聚合、销量矩阵、样品统计）
"""

import glob
import math
import os
from datetime import datetime, timedelta
from urllib.parse import quote

import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def find_project_root(start_dir):
    """向上查找 sum_daily_order 根目录，避免依赖固定绝对路径。"""
    env_root = os.getenv("TIKTOK_REPORT_ROOT")
    if env_root:
        root = os.path.abspath(env_root)
        for name in ("config", "data", "result"):
            os.makedirs(os.path.join(root, name), exist_ok=True)
        return root

    current = os.path.abspath(start_dir)
    while True:
        if all(os.path.isdir(os.path.join(current, name)) for name in ("config", "data", "result")):
            return current
        parent = os.path.dirname(current)
        if parent == current:
            raise FileNotFoundError("无法定位 sum_daily_order 根目录，请确认 config/data/result 目录仍在同一层级。")
        current = parent


# ============================================================
# 0. 基础配置
# ============================================================
CURRENT_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = find_project_root(CURRENT_SCRIPT_DIR)

for env_path in [
    os.path.join(PROJECT_ROOT, "config", ".env"),
    os.path.join(CURRENT_SCRIPT_DIR, ".env"),
]:
    if os.path.exists(env_path):
        load_dotenv(env_path)
load_dotenv()

FEISHU_APP_ID = os.getenv("FEISHU_APP_ID")
FEISHU_APP_SECRET = os.getenv("FEISHU_APP_SECRET")

FEISHU_SHEET_TOKEN = "G3HCsMq7UhSjIptrTI5c0dUnnyh"
FEISHU_RANGE_SKU = "A:G"
FEISHU_REQUIRED_SCOPE_HINT = "请在飞书开放平台给应用开通 sheets:spreadsheet:readonly 或 sheets:spreadsheet:read 权限，并重新发布/生效。"

# -------- 越南固定汇率（1 越南盾兑人民币）--------
VIETNAM_EXCHANGE_RATE = 1 / 3883

SKU_ID_COLUMN = "SKU ID"
PRODUCT_CATEGORY_COLUMN = "产品大类"
LOGISTICS_CAPACITY_COLUMN = "每单物流承载数量"
CACHE_DIR = os.path.join(PROJECT_ROOT, "config", "cache")
os.makedirs(CACHE_DIR, exist_ok=True)
CACHE_EXPIRY_HOURS = 24
USE_CONFIG_CACHE_FIRST = os.getenv("USE_CONFIG_CACHE_FIRST", "").lower() in {"1", "true", "yes", "y"}
SHEET_ID_CACHE = {}

# 如果某个店铺暂时不需要运行，把 enabled 改成 False 即可。
VIETNAM_STORES = [
    {
        "enabled": True,
        "country_code": "VN",
        "country_name": "越南",
        "store_key": "local",
        "store_name": "本土店",
        "store_dir": "local",
        "sheet_name": "越南_本土店",
    },
    {
        "enabled": True,
        "country_code": "VN",
        "country_name": "越南",
        "store_key": "cross_border",
        "store_name": "跨境店",
        "store_dir": "cross_border",
        "sheet_name": "越南_跨境店",
    },
]


# ============================================================
# 1. 飞书表格读取 + 缓存工具
# ============================================================
def get_feishu_token(app_id, app_secret):
    """获取飞书 tenant_access_token。"""
    url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
    payload = {"app_id": app_id, "app_secret": app_secret}
    try:
        resp = requests.post(url, json=payload, timeout=10)
        if resp.status_code != 200:
            raise Exception(f"飞书Token请求失败，HTTP {resp.status_code}")
        data = resp.json()
        if data.get("code") != 0:
            raise Exception(f"飞书Token错误: {data}")
        return data["tenant_access_token"]
    except Exception as e:
        raise Exception(f"获取飞书 token 失败: {e}")


def format_feishu_error(data):
    """把飞书 API 错误整理成控制台可读信息。"""
    if not isinstance(data, dict):
        return str(data)

    code = data.get("code")
    msg = data.get("msg")
    error = data.get("error") or {}
    message = error.get("message")
    log_id = error.get("log_id")

    parts = [f"code={code}", f"msg={msg}"]
    if message:
        parts.append(f"message={message}")
    if log_id:
        parts.append(f"log_id={log_id}")
    if code == 99991672:
        parts.append(FEISHU_REQUIRED_SCOPE_HINT)
    return "；".join(parts)


def get_feishu_json(url, headers, action):
    """请求飞书接口并保留详细错误。"""
    resp = requests.get(url, headers=headers, timeout=10)
    try:
        data = resp.json()
    except Exception:
        data = {"code": resp.status_code, "msg": resp.text[:500]}

    if resp.status_code != 200 or data.get("code") != 0:
        raise Exception(f"{action}失败，HTTP {resp.status_code}，{format_feishu_error(data)}")
    return data


def get_feishu_sheet_id(token, sheet_token, sheet_name):
    """飞书 values API 需要 sheet_id，这里支持用工作表名称自动解析。"""
    cache_key = (sheet_token, sheet_name)
    if cache_key in SHEET_ID_CACHE:
        return SHEET_ID_CACHE[cache_key]

    headers = {"Authorization": f"Bearer {token}"}
    url = f"https://open.feishu.cn/open-apis/sheets/v3/spreadsheets/{sheet_token}/sheets/query"
    data = get_feishu_json(url, headers, "查询飞书工作表列表")
    sheets = data.get("data", {}).get("sheets", [])

    available = []
    for item in sheets:
        props = item.get("sheet") or item.get("properties") or item
        title = props.get("title") or props.get("sheet_name")
        sheet_id = props.get("sheet_id") or props.get("sheetId")
        if title and sheet_id:
            SHEET_ID_CACHE[(sheet_token, title)] = sheet_id
            SHEET_ID_CACHE[(sheet_token, sheet_id)] = sheet_id
            available.append(f"{title}({sheet_id})")
            if title == sheet_name or sheet_id == sheet_name:
                return sheet_id

    raise ValueError(f"未找到飞书 Sheet '{sheet_name}'。当前可用 Sheet: {', '.join(available) or '无'}")


def read_feishu_sheet(token, sheet_token, sheet_name, range_str):
    """读取飞书表格指定 Sheet 的数据，返回 DataFrame（第一行为列名）。"""
    headers = {"Authorization": f"Bearer {token}"}
    sheet_id = get_feishu_sheet_id(token, sheet_token, sheet_name)
    range_expr = quote(f"{sheet_id}!{range_str}", safe="")
    url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{sheet_token}/values/{range_expr}"
    try:
        data = get_feishu_json(url, headers, "读取飞书表格")
        values = data.get("data", {}).get("valueRange", {}).get("values", [])
        if not values:
            raise ValueError(f"飞书表格 Sheet '{sheet_name}' 返回空数据")
        return pd.DataFrame(values[1:], columns=values[0])
    except Exception as e:
        raise Exception(f"读取飞书表格 Sheet '{sheet_name}' 失败: {e}")


def clean_sku_id_column(df):
    """统一清洗配置表 SKU ID。"""
    if SKU_ID_COLUMN not in df.columns:
        return df
    df = df.dropna(subset=[SKU_ID_COLUMN]).copy()
    df[SKU_ID_COLUMN] = df[SKU_ID_COLUMN].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    df = df[df[SKU_ID_COLUMN] != ""]
    return df


def clean_config_dataframe(df):
    """统一清洗配置表：SKU 转字符串，成本列转数字。"""
    df = clean_sku_id_column(df)
    cost_cols = ["产品成本(元)", "物流成本(元)", "寄样成本(元)"]
    raw_sample_cost = df["寄样成本(元)"].copy() if "寄样成本(元)" in df.columns else None

    for col in cost_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    if "寄样成本(元)" not in df.columns and {"产品成本(元)", "物流成本(元)"}.issubset(df.columns):
        df["寄样成本(元)"] = df["产品成本(元)"] + df["物流成本(元)"]
    if {"产品成本(元)", "物流成本(元)", "寄样成本(元)"}.issubset(df.columns):
        empty_sample_cost = df["寄样成本(元)"].isna() | (df["寄样成本(元)"] <= 0)
        df.loc[empty_sample_cost, "寄样成本(元)"] = (
            df.loc[empty_sample_cost, "产品成本(元)"] + df.loc[empty_sample_cost, "物流成本(元)"]
        )
    if LOGISTICS_CAPACITY_COLUMN not in df.columns:
        df[LOGISTICS_CAPACITY_COLUMN] = 1
    df[LOGISTICS_CAPACITY_COLUMN] = pd.to_numeric(df[LOGISTICS_CAPACITY_COLUMN], errors="coerce").fillna(1)
    df.loc[df[LOGISTICS_CAPACITY_COLUMN] <= 0, LOGISTICS_CAPACITY_COLUMN] = 1

    if raw_sample_cost is not None:
        formula_mask = raw_sample_cost.astype(str).str.match(r"^=?C\d+\+D\d+$", na=False)
        if formula_mask.any() and set(cost_cols).issubset(df.columns):
            df.loc[formula_mask, "寄样成本(元)"] = (
                df.loc[formula_mask, "产品成本(元)"] + df.loc[formula_mask, "物流成本(元)"]
            )
    if PRODUCT_CATEGORY_COLUMN not in df.columns:
        df[PRODUCT_CATEGORY_COLUMN] = df.get("中文简称", "")
    for col in ["中文简称", PRODUCT_CATEGORY_COLUMN]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    if "中文简称" in df.columns:
        df[PRODUCT_CATEGORY_COLUMN] = df[PRODUCT_CATEGORY_COLUMN].replace({"": pd.NA, "nan": pd.NA})
        df[PRODUCT_CATEGORY_COLUMN] = df[PRODUCT_CATEGORY_COLUMN].fillna(df["中文简称"])
    return df


def get_cache_file_path(sheet_name):
    safe_name = sheet_name.replace(" ", "_").replace("/", "_")
    return os.path.join(CACHE_DIR, f"config_{safe_name}.csv")


def get_config_dataframe(sheet_name, force_refresh=False):
    """获取指定 Sheet 的配置 DataFrame，优先飞书，失败则用本地缓存。"""
    cache_file = get_cache_file_path(sheet_name)
    cache_valid = False
    if os.path.exists(cache_file) and not force_refresh:
        mtime = datetime.fromtimestamp(os.path.getmtime(cache_file))
        if datetime.now() - mtime < timedelta(hours=CACHE_EXPIRY_HOURS):
            cache_valid = True

    if cache_valid and USE_CONFIG_CACHE_FIRST and not force_refresh:
        try:
            df = pd.read_csv(cache_file, dtype=str, encoding="utf-8-sig")
            df = clean_config_dataframe(df)
            print(f"   ✅ 使用本地缓存配置（Sheet: {sheet_name}）")
            return df
        except Exception as e:
            print(f"   ⚠️ 缓存读取失败: {e}，尝试刷新")

    try:
        if not FEISHU_APP_ID or not FEISHU_APP_SECRET:
            raise EnvironmentError(
                "未配置 FEISHU_APP_ID/FEISHU_APP_SECRET，无法从飞书刷新配置。"
            )
        token = get_feishu_token(FEISHU_APP_ID, FEISHU_APP_SECRET)
        df = read_feishu_sheet(token, FEISHU_SHEET_TOKEN, sheet_name, FEISHU_RANGE_SKU)
        df = df.dropna(how="all")
        df = clean_config_dataframe(df)
        df.to_csv(cache_file, index=False, encoding="utf-8-sig")
        print(f"   ✅ 成功从飞书拉取配置并缓存（Sheet: {sheet_name}）")
        return df
    except Exception as e:
        print(f"   ❌ 飞书拉取失败（Sheet: {sheet_name}）: {e}")
        if os.path.exists(cache_file):
            try:
                df = pd.read_csv(cache_file, dtype=str, encoding="utf-8-sig")
                df = clean_config_dataframe(df)
                print(f"   ⚠️ 飞书不可用，使用过期缓存（Sheet: {sheet_name}）")
                return df
            except Exception:
                pass
        raise Exception(f"无法获取配置数据（Sheet: {sheet_name}）")


# ============================================================
# 2. 辅助函数
# ============================================================
def parse_amount(value):
    """将字符串金额转为浮点数。"""
    if pd.isna(value):
        return 0.0
    s = (
        str(value)
        .replace(" ", "")
        .replace("\t", "")
        .replace(",", "")
        .replace("VND", "")
        .replace("₫", "")
    )
    try:
        return float(s)
    except Exception:
        return 0.0


def clean_sku_str(sku_series):
    """将 Series 中的 SKU 转换为字符串并去除 .0 后缀和空格。"""
    return sku_series.astype(str).str.replace(r"\.0$", "", regex=True).str.strip()


def parse_order_date(value):
    """越南订单时间通常是日/月/年格式，统一输出为 YYYY-MM-DD。"""
    if pd.isna(value):
        return pd.NA
    s = str(value).strip()
    if not s:
        return pd.NA
    try:
        return pd.to_datetime(s, dayfirst=True).strftime("%Y-%m-%d")
    except Exception:
        return pd.NA


def validate_config_columns(config_df, sheet_name):
    required_cols = [SKU_ID_COLUMN, "中文简称", "产品成本(元)", "物流成本(元)", "寄样成本(元)"]
    missing = [col for col in required_cols if col not in config_df.columns]
    if missing:
        raise ValueError(f"配置表 {sheet_name} 缺少列: {', '.join(missing)}")
    if config_df.empty:
        raise ValueError(f"配置表 {sheet_name} 没有 SKU 数据，请先补充 SKU ID、中文简称和成本信息")


def safe_change_rate(current, previous):
    """计算环比变化率；上一期为 0 时返回空值，避免误导。"""
    if previous == 0:
        return ""
    return round((current - previous) / previous, 4)


def format_percent(value):
    """把小数格式化为百分比文本，便于 Excel 中直接阅读。"""
    if value == "" or pd.isna(value):
        return ""
    return f"{value:.2%}"

def build_order_logistics_allocation(df_normal, logistics_col):
    """按订单计算物流成本，并分摊到订单内的产品规格。"""
    result_cols = ["日期", "Product Category", "Mapped Name", "物流成本"]
    if df_normal.empty:
        return pd.DataFrame(columns=result_cols)

    df_calc = df_normal.copy()
    if "Order ID" in df_calc.columns:
        order_key = df_calc["Order ID"].astype(str).str.strip()
        missing_order = order_key.isna() | order_key.eq("") | order_key.eq("nan")
        order_key = order_key.mask(missing_order, "ROW_" + df_calc.index.astype(str))
    else:
        order_key = "ROW_" + df_calc.index.astype(str)
    df_calc["_OrderKey"] = order_key
    df_calc[LOGISTICS_CAPACITY_COLUMN] = pd.to_numeric(
        df_calc.get(LOGISTICS_CAPACITY_COLUMN, 1),
        errors="coerce",
    ).fillna(1)
    df_calc.loc[df_calc[LOGISTICS_CAPACITY_COLUMN] <= 0, LOGISTICS_CAPACITY_COLUMN] = 1

    category_order = df_calc.groupby(["_OrderKey", "日期", "Product Category"]).agg(
        category_qty=("SKU_ID", "count"),
        original_logistics=(logistics_col, "sum"),
        unit_logistics=(logistics_col, "max"),
        capacity=(LOGISTICS_CAPACITY_COLUMN, "max"),
    ).reset_index()
    category_order["_capacity_logistics"] = category_order.apply(
        lambda row: math.ceil(row["category_qty"] / row["capacity"]) * row["unit_logistics"],
        axis=1,
    )
    category_order["candidate_logistics"] = category_order[["original_logistics", "_capacity_logistics"]].min(axis=1)
    order_max = category_order.groupby("_OrderKey")["candidate_logistics"].transform("max")
    order_sum = category_order.groupby("_OrderKey")["candidate_logistics"].transform("sum")
    category_order["category_logistics"] = category_order.apply(
        lambda row: 0 if order_sum.loc[row.name] == 0 else order_max.loc[row.name] * row["candidate_logistics"] / order_sum.loc[row.name],
        axis=1,
    )

    sku_order = df_calc.groupby(["_OrderKey", "日期", "Product Category", "Mapped Name"]).agg(
        sku_qty=("SKU_ID", "count"),
        sku_original_logistics=(logistics_col, "sum"),
    ).reset_index()
    sku_order = sku_order.merge(
        category_order[["_OrderKey", "日期", "Product Category", "category_qty", "original_logistics", "category_logistics"]],
        on=["_OrderKey", "日期", "Product Category"],
        how="left",
    )
    sku_order["物流成本"] = sku_order.apply(
        lambda row: (
            0
            if row["category_qty"] == 0
            else row["category_logistics"] * (
                row["sku_original_logistics"] / row["original_logistics"]
                if row["original_logistics"] != 0
                else row["sku_qty"] / row["category_qty"]
            )
        ),
        axis=1,
    )
    return sku_order.groupby(["日期", "Product Category", "Mapped Name"])["物流成本"].sum().reset_index()


def build_period_comparison_frames(df_daily_product, daily_order_records):
    """按 CSV 文件生成周期汇总；多文件时追加相邻周期对比。"""
    if df_daily_product.empty or "文件名" not in df_daily_product.columns:
        return []

    period_summary = df_daily_product.groupby("文件名").agg({
        "销量": "sum",
        "寄样数": "sum",
        "销售额": "sum",
        "汇率后金额": "sum",
        "P列折后价": "sum",
        "产品成本": "sum",
        "物流成本": "sum",
        "寄样支出": "sum",
        "日期": ["min", "max"],
    })
    period_summary.columns = [
        "销量", "寄样数", "销售额", "汇率后金额", "P列折后价",
        "产品成本", "物流成本", "寄样支出", "开始日期", "结束日期"
    ]

    if daily_order_records:
        df_order_count = (
            pd.DataFrame(daily_order_records)
            .drop_duplicates(subset=["文件名", "日期", "Order ID"])
            .groupby("文件名")["Order ID"]
            .count()
            .rename("订单数")
        )
        period_summary = period_summary.join(df_order_count, how="left")
    else:
        period_summary["订单数"] = period_summary["销量"]

    period_summary["订单数"] = period_summary["订单数"].fillna(0)
    period_summary["利润"] = (
        period_summary["汇率后金额"]
        - period_summary["产品成本"]
        - period_summary["物流成本"]
        - period_summary["寄样支出"]
    )
    period_summary["利润率"] = period_summary.apply(
        lambda row: 0 if row["汇率后金额"] == 0 else row["利润"] / row["汇率后金额"],
        axis=1,
    )
    period_summary = period_summary.reset_index()
    period_summary["_开始日期排序"] = pd.to_datetime(period_summary["开始日期"], errors="coerce")
    period_summary["_结束日期排序"] = pd.to_datetime(period_summary["结束日期"], errors="coerce")
    period_summary = (
        period_summary
        .sort_values(["_开始日期排序", "_结束日期排序", "文件名"])
        .drop(columns=["_开始日期排序", "_结束日期排序"])
        .set_index("文件名")
    )

    metrics = ["订单数", "销售额", "汇率后金额", "P列折后价", "产品成本", "物流成本", "寄样支出", "利润", "利润率", "销量", "寄样数"]
    frames = []
    ordered_files = list(period_summary.index)

    summary_rows = []
    for file_name in ordered_files:
        row = {"文件名": file_name}
        for metric in metrics:
            if metric == "利润率":
                row[metric] = format_percent(period_summary.loc[file_name, metric])
            else:
                row[metric] = round(period_summary.loc[file_name, metric], 2)
        summary_rows.append(row)
    frames.append(("周期汇总", pd.DataFrame(summary_rows)))

    for idx in range(1, len(ordered_files)):
        previous_file = ordered_files[idx - 1]
        current_file = ordered_files[idx]
        previous = period_summary.loc[previous_file]
        current = period_summary.loc[current_file]
        change_value_row = {"文件名": f"变化值：{current_file} - {previous_file}"}
        change_rate_row = {"文件名": f"变化率：{current_file} / {previous_file} - 1"}
        for metric in metrics:
            if metric == "利润率":
                change_value_row[metric] = format_percent(current[metric] - previous[metric])
                change_rate_row[metric] = ""
            else:
                previous_value = round(previous[metric], 2)
                current_value = round(current[metric], 2)
                change_value_row[metric] = round(current_value - previous_value, 2)
                change_rate_row[metric] = format_percent(safe_change_rate(current_value, previous_value))
        title = f"{current_file} - {previous_file}"
        frames.append((title, pd.DataFrame([change_value_row, change_rate_row])))
    return frames

def center_excel_sheet(writer, sheet_name, row_count, col_count):
    """把已写入的工作表区域设置为居中显示，并冻结首行。"""
    worksheet = writer.sheets[sheet_name]
    if hasattr(worksheet, "set_column"):
        worksheet.freeze_panes(1, 0)
        if not hasattr(writer, "_center_format"):
            writer._center_format = writer.book.add_format({"align": "center", "valign": "vcenter"})
        center_format = writer._center_format
        worksheet.set_column(0, max(col_count - 1, 0), 14, center_format)
        for row_idx in range(row_count):
            worksheet.set_row(row_idx, None, center_format)
        return

    worksheet.freeze_panes = "A2"
    center_alignment = Alignment(horizontal="center", vertical="center")
    for row in worksheet.iter_rows(min_row=1, max_row=row_count, min_col=1, max_col=col_count):
        for cell in row:
            cell.alignment = center_alignment

    for col_idx in range(1, col_count + 1):
        max_width = 10
        for row_idx in range(1, row_count + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            text = str(value)
            width = sum(2 if ord(char) > 127 else 1 for char in text)
            max_width = max(max_width, width + 2)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_width, 28)

    for row_idx in range(1, row_count + 1):
        worksheet.row_dimensions[row_idx].height = 20


def insert_blank_rows_between_files(df):
    """输出展示用：不同文件名之间插入空行，不参与任何计算。"""
    if df.empty or "文件名" not in df.columns:
        return df

    summary_labels = {"全部文件合计", "汇总", "总计"}
    parts = []
    blank_row = {col: "" for col in df.columns}
    for file_name, group in df.groupby("文件名", sort=False, dropna=False):
        if str(file_name).strip() in summary_labels:
            parts.append(group)
            continue
        if parts:
            parts.append(pd.DataFrame([blank_row], columns=df.columns))
        parts.append(group)
    return pd.concat(parts, ignore_index=True)


def build_product_profit_by_period(df_daily_product):
    """按文件+产品汇总 GMV、成本和利润，方便判断单品盈利情况。"""
    if df_daily_product.empty:
        return pd.DataFrame([["无产品汇总数据"]], columns=["提示"])

    summary = df_daily_product.groupby(["文件名", "产品名称"]).agg({
        "销量": "sum",
        "寄样数": "sum",
        "销售额": "sum",
        "汇率后金额": "sum",
        "P列折后价": "sum",
        "产品成本": "sum",
        "物流成本": "sum",
        "寄样支出": "sum",
    }).reset_index()
    summary["总成本"] = summary["产品成本"] + summary["物流成本"] + summary["寄样支出"]
    summary["利润"] = summary["汇率后金额"] - summary["总成本"]
    summary["利润率"] = summary.apply(
        lambda row: "" if row["汇率后金额"] == 0 else format_percent(row["利润"] / row["汇率后金额"]),
        axis=1,
    )

    numeric_cols = ["销售额", "汇率后金额", "P列折后价", "产品成本", "物流成本", "寄样支出", "总成本", "利润"]
    for col in numeric_cols:
        summary[col] = summary[col].round(2)
    return summary[[
        "文件名", "产品名称", "销售额", "汇率后金额", "P列折后价",
        "产品成本", "物流成本", "寄样支出", "总成本", "利润", "利润率"
    ]]


def build_sku_profit_by_period(df_sku_detail):
    """按文件+产品大类+SKU 汇总，放在 Daily Product Detail 下半部分。"""
    if df_sku_detail.empty:
        return pd.DataFrame([["无 SKU 汇总数据"]], columns=["提示"])

    summary = df_sku_detail.groupby(["文件名", "产品大类", "产品名称"]).agg({
        "销量": "sum",
        "寄样数": "sum",
        "销售额": "sum",
        "汇率后金额": "sum",
        "P列折后价": "sum",
        "产品成本": "sum",
        "物流成本": "sum",
        "寄样支出": "sum",
    }).reset_index()
    summary["总成本"] = summary["产品成本"] + summary["物流成本"] + summary["寄样支出"]
    summary["利润"] = summary["汇率后金额"] - summary["总成本"]
    summary["利润率"] = summary.apply(
        lambda row: "" if row["汇率后金额"] == 0 else format_percent(row["利润"] / row["汇率后金额"]),
        axis=1,
    )

    numeric_cols = ["销售额", "汇率后金额", "P列折后价", "产品成本", "物流成本", "寄样支出", "总成本", "利润"]
    for col in numeric_cols:
        summary[col] = summary[col].round(2)
    return summary[[
        "文件名", "产品大类", "产品名称", "销售额", "汇率后金额", "P列折后价",
        "产品成本", "物流成本", "寄样支出", "总成本", "利润", "利润率", "销量", "寄样数"
    ]]


def build_product_quantity_by_period(product_quantity_records):
    """按文件汇总每个产品的销量，放在每日销量矩阵上方。"""
    df_quantity_records = pd.DataFrame(product_quantity_records)
    if df_quantity_records.empty:
        return pd.DataFrame([["无销量数据"]], columns=["提示"])

    matrix = df_quantity_records.pivot_table(
        index="产品名称",
        columns="文件名",
        values="销量",
        aggfunc="sum",
        fill_value=0,
    ).astype(int)
    matrix["汇总"] = matrix.sum(axis=1)
    matrix.loc["汇总"] = matrix.sum(axis=0)
    return matrix


def build_daily_product_quantity_matrix(df_daily_product):
    """按日期横向展开每个产品的销量，用于 Daily Product Detail。"""
    if df_daily_product.empty:
        return pd.DataFrame([["无每日产品销量数据"]], columns=["提示"])

    matrix = df_daily_product.pivot_table(
        index="产品名称",
        columns="日期",
        values="销量",
        aggfunc="sum",
        fill_value=0,
    ).astype(int)
    sorted_cols = sorted(matrix.columns, key=lambda x: pd.to_datetime(x, errors="coerce"))
    matrix = matrix[sorted_cols]
    matrix["汇总"] = matrix.sum(axis=1)
    matrix.loc["汇总"] = matrix.sum(axis=0)
    return matrix.reset_index()


def collect_order_sku_preview(store_config, max_rows=80):
    """配置表为空时，输出订单里出现过的 SKU，方便补配置。"""
    country_code = store_config["country_code"]
    store_dir = store_config["store_dir"]
    folder_path = os.path.normpath(os.path.join(PROJECT_ROOT, "data", f"data_{country_code}", store_dir))
    file_list = glob.glob(os.path.join(folder_path, "*.csv"))

    preview = {}
    for file_path in file_list:
        try:
            df = pd.read_csv(file_path, dtype=str)
            df.columns = df.columns.astype(str).str.strip()
        except Exception:
            continue

        if SKU_ID_COLUMN not in df.columns:
            continue

        for _, row in df.iterrows():
            sku_id = str(row.get(SKU_ID_COLUMN, "")).replace(".0", "").strip()
            if not sku_id or sku_id.lower() == "nan" or sku_id in preview:
                continue

            preview[sku_id] = {
                "Seller SKU": str(row.get("Seller SKU", "")).strip(),
                "Product Name": str(row.get("Product Name", "")).strip(),
            }
            if len(preview) >= max_rows:
                return preview
    return preview


def print_order_sku_preview(store_config):
    preview = collect_order_sku_preview(store_config)
    if not preview:
        return

    display_name = f"{store_config['country_name']}_{store_config['store_name']}"
    print(f"   💡 [{display_name}] 订单中出现的 SKU ID，可用于补充配置表:")
    for sku_id, info in preview.items():
        seller_sku = info["Seller SKU"]
        product_name = info["Product Name"]
        if len(product_name) > 60:
            product_name = product_name[:60] + "..."
        print(f"      - {sku_id} | {seller_sku} | {product_name}")


# ============================================================
# 3. 核心处理函数
# ============================================================
def run_report(store_config, config_df, exchange_rate):
    country_code = store_config["country_code"]
    country_name = store_config["country_name"]
    store_key = store_config["store_key"]
    store_name = store_config["store_name"]
    store_dir = store_config["store_dir"]
    display_name = f"{country_name}_{store_name}"

    validate_config_columns(config_df, store_config["sheet_name"])

    folder_path = os.path.normpath(os.path.join(PROJECT_ROOT, "data", f"data_{country_code}", store_dir))
    output_filename = os.path.normpath(
        os.path.join(PROJECT_ROOT, f"result/Daily_Performance_Report_{country_code.upper()}_{store_key}.xlsx")
    )

    file_list = glob.glob(os.path.join(folder_path, "*.csv"))
    if not file_list:
        print(f"⚠️  [{display_name}] 路径下未找到文件: {folder_path}")
        return

    print(f"🚀 开始处理 [{display_name}]，共 {len(file_list)} 个文件")

    status_col = "Order Status"
    quantity_col = "Normal or Pre-order"
    n_col = "SKU Platform Discount"
    p_col = "SKU Subtotal After Discount"
    r_col = "Original Shipping Fee"
    logistics_col = "物流成本(元)"
    sample_cost_col = "寄样成本(元)"
    cancel_keywords = ["Canceled", "Cancelled", "Unpaid", "已取消", "未支付", "未付款", "部分发货后取消"]

    daily_product_detail_list = []
    sku_order_detail_list = []
    product_quantity_records = []
    sample_quantity_records = []
    daily_order_records = []
    unmatched_skus = set()

    for file_path in file_list:
        file_name = os.path.basename(file_path)
        try:
            df = pd.read_csv(file_path)
            df.columns = df.columns.astype(str).str.strip()
        except Exception as e:
            print(f"读取失败 {file_name}: {e}")
            continue

        required_order_cols = [status_col, quantity_col, SKU_ID_COLUMN, n_col, p_col, r_col]
        missing_order_cols = [col for col in required_order_cols if col not in df.columns]
        if missing_order_cols:
            print(f"⚠️ 文件 {file_name} 缺少必要列: {', '.join(missing_order_cols)}，跳过")
            continue

        status_series = df[status_col].astype(str).str.strip()
        quantity_series = df[quantity_col].astype(str).str.strip()
        normal_mask = ~status_series.isin(cancel_keywords) & quantity_series.eq("Normal")

        # -------- 正常订单 --------
        df_normal = df[normal_mask].copy()
        if "Paid Time" not in df_normal.columns:
            print(f"⚠️ 文件 {file_name} 中未找到 Paid Time 列，将跳过此文件")
            continue
        df_normal["日期"] = df_normal["Paid Time"].map(parse_order_date)
        df_normal = df_normal.dropna(subset=["日期"]).copy()

        for col in [n_col, p_col, r_col]:
            df_normal[col] = df_normal[col].map(parse_amount)
        df_normal["SKU_ID"] = clean_sku_str(df_normal[SKU_ID_COLUMN])

        df_normal = df_normal.merge(
            config_df[[SKU_ID_COLUMN, PRODUCT_CATEGORY_COLUMN, "中文简称", "产品成本(元)", logistics_col, sample_cost_col, LOGISTICS_CAPACITY_COLUMN]],
            left_on="SKU_ID",
            right_on=SKU_ID_COLUMN,
            how="left",
        )

        missing = df_normal[df_normal["中文简称"].isna()]
        if not missing.empty:
            for sku in missing["SKU_ID"].unique():
                unmatched_skus.add(sku)

        df_normal["Mapped Name"] = df_normal["中文简称"].fillna(df_normal["SKU_ID"])
        df_normal["Product Category"] = df_normal[PRODUCT_CATEGORY_COLUMN].fillna(df_normal["Mapped Name"])
        if "Order ID" in df_normal.columns:
            daily_order_records.extend(
                df_normal[["日期", "Order ID"]]
                .dropna()
                .drop_duplicates()
                .assign(文件名=file_name)
                .to_dict("records")
            )

        norm_agg = df_normal.groupby(["日期", "Product Category", "Mapped Name"]).agg({
            n_col: "sum",
            p_col: "sum",
            r_col: "sum",
            "产品成本(元)": "first",
            sample_cost_col: "first",
            "SKU_ID": "count",
        }).reset_index()
        norm_agg.rename(columns={"SKU_ID": "销量"}, inplace=True)
        if "Order ID" in df_normal.columns:
            sku_daily_order_count = (
                df_normal[["日期", "Product Category", "Mapped Name", "Order ID"]]
                .dropna(subset=["Order ID"])
                .assign(Order_ID_Clean=lambda x: x["Order ID"].astype(str).str.strip())
            )
            sku_daily_order_count = (
                sku_daily_order_count[sku_daily_order_count["Order_ID_Clean"] != ""]
                .groupby(["日期", "Product Category", "Mapped Name"])["Order_ID_Clean"]
                .nunique()
                .reset_index()
                .rename(columns={"Order_ID_Clean": "订单数"})
            )
        else:
            sku_daily_order_count = norm_agg[["日期", "Product Category", "Mapped Name", "销量"]].copy()
            sku_daily_order_count.rename(columns={"销量": "订单数"}, inplace=True)
        norm_agg["销售额"] = norm_agg[n_col] + norm_agg[p_col] + norm_agg[r_col]
        norm_agg["产品成本"] = norm_agg["销量"] * norm_agg["产品成本(元)"]
        logistics_daily_product = build_order_logistics_allocation(df_normal, logistics_col)
        norm_agg = norm_agg.merge(
            logistics_daily_product,
            on=["日期", "Product Category", "Mapped Name"],
            how="left",
        )
        norm_agg["物流成本"] = norm_agg["物流成本"].fillna(0)

        # -------- 样品订单 --------
        sample_mask = ~status_series.isin(cancel_keywords) & (
            df[quantity_col].isna() | quantity_series.eq("") | quantity_series.eq("nan")
        )
        df_sample = df[sample_mask].copy()
        if not df_sample.empty:
            if "Paid Time" not in df_sample.columns:
                print(f"⚠️ 文件 {file_name} 样品订单中未找到 Paid Time 列，将跳过样品统计")
                df_sample = pd.DataFrame()
            else:
                df_sample["日期"] = df_sample["Paid Time"].map(parse_order_date)
                df_sample = df_sample.dropna(subset=["日期"]).copy()

        if not df_sample.empty:
            df_sample["SKU_ID"] = clean_sku_str(df_sample[SKU_ID_COLUMN])
            df_sample = df_sample.merge(
                config_df[[SKU_ID_COLUMN, PRODUCT_CATEGORY_COLUMN, "中文简称", sample_cost_col]],
                left_on="SKU_ID",
                right_on=SKU_ID_COLUMN,
                how="left",
            )
            missing_s = df_sample[df_sample["中文简称"].isna()]
            if not missing_s.empty:
                for sku in missing_s["SKU_ID"].unique():
                    unmatched_skus.add(sku)
            df_sample["Mapped Name"] = df_sample["中文简称"].fillna(df_sample["SKU_ID"])
            df_sample["Product Category"] = df_sample[PRODUCT_CATEGORY_COLUMN].fillna(df_sample["Mapped Name"])

            sample_daily_product = df_sample.groupby(["日期", "Product Category", "Mapped Name"]).agg({
                sample_cost_col: "first",
                "SKU_ID": "count",
            }).reset_index()
            sample_daily_product.rename(columns={"SKU_ID": "寄样数"}, inplace=True)
            sample_daily_product["寄样支出"] = sample_daily_product["寄样数"] * sample_daily_product[sample_cost_col]
            sample_daily_product = sample_daily_product[["日期", "Product Category", "Mapped Name", "寄样数", "寄样支出"]]
        else:
            sample_daily_product = pd.DataFrame(columns=["日期", "Product Category", "Mapped Name", "寄样数", "寄样支出"])

        merged_daily_product = norm_agg[[
            "日期", "Product Category", "Mapped Name", "销量", "销售额", p_col, "产品成本", "物流成本"
        ]].merge(
            sample_daily_product,
            on=["日期", "Product Category", "Mapped Name"],
            how="outer",
        )
        for col in ["销量", "销售额", p_col, "产品成本", "物流成本", "寄样数", "寄样支出"]:
            if col in merged_daily_product.columns:
                merged_daily_product[col] = pd.to_numeric(merged_daily_product[col], errors="coerce").fillna(0)
        merged_daily_product["汇率后金额"] = (merged_daily_product["销售额"] * exchange_rate).round(2)

        for _, row in merged_daily_product.iterrows():
            daily_product_detail_list.append({
                "文件名": file_name,
                "日期": row["日期"],
                "产品大类": row["Product Category"],
                "产品名称": row["Mapped Name"],
                "销量": int(row["销量"]),
                "寄样数": int(row["寄样数"]),
                "销售额": round(row["销售额"], 2),
                "汇率后金额": round(row["汇率后金额"], 2),
                "P列折后价": round(row[p_col], 2),
                "产品成本": round(row["产品成本"], 2),
                "物流成本": round(row["物流成本"], 2),
                "寄样支出": round(row["寄样支出"], 2),
            })

        for _, row in sku_daily_order_count.iterrows():
            sku_order_detail_list.append({
                "文件名": file_name,
                "日期": row["日期"],
                "产品大类": row["Product Category"],
                "产品名称": row["Mapped Name"],
                "订单数": int(row["订单数"]),
            })

        for _, row in norm_agg.iterrows():
            product_quantity_records.append({
                "文件名": file_name,
                "日期": row["日期"],
                "产品名称": row["Product Category"],
                "销量": int(row["销量"]),
            })

        for _, row in sample_daily_product.iterrows():
            sample_quantity_records.append({
                "文件名": file_name,
                "日期": row["日期"],
                "产品名称": row["Product Category"],
                "寄样数": int(row["寄样数"]),
            })

    if not daily_product_detail_list:
        print(f"⚠️ [{display_name}] 没有有效数据，不生成报表。")
        return

    df_sku_detail = pd.DataFrame(daily_product_detail_list)
    if sku_order_detail_list:
        df_sku_order_count = pd.DataFrame(sku_order_detail_list) \
            .groupby(["文件名", "日期", "产品大类", "产品名称"])["订单数"] \
            .sum().reset_index()
        df_sku_detail = df_sku_detail.merge(
            df_sku_order_count,
            on=["文件名", "日期", "产品大类", "产品名称"],
            how="left",
        )
    else:
        df_sku_detail["订单数"] = 0
    df_sku_detail["订单数"] = df_sku_detail["订单数"].fillna(0).astype(int)
    df_daily_product = df_sku_detail.groupby(["文件名", "日期", "产品大类"]).agg({
        "销量": "sum",
        "寄样数": "sum",
        "销售额": "sum",
        "汇率后金额": "sum",
        "P列折后价": "sum",
        "产品成本": "sum",
        "物流成本": "sum",
        "寄样支出": "sum",
    }).reset_index().rename(columns={"产品大类": "产品名称"})
    df_daily_summary = df_daily_product.groupby(["文件名", "日期"]).agg({
        "销量": "sum",
        "寄样数": "sum",
        "销售额": "sum",
        "汇率后金额": "sum",
        "P列折后价": "sum",
        "产品成本": "sum",
        "物流成本": "sum",
        "寄样支出": "sum",
    }).reset_index()
    if daily_order_records:
        df_order_count = pd.DataFrame(daily_order_records) \
            .drop_duplicates(subset=["文件名", "日期", "Order ID"]) \
            .groupby(["文件名", "日期"])["Order ID"].count().reset_index()
        df_order_count.rename(columns={"Order ID": "订单数"}, inplace=True)
    else:
        df_order_count = df_daily_product[df_daily_product["销量"] > 0].groupby(["文件名", "日期"])["销量"].sum().reset_index()
        df_order_count.rename(columns={"销量": "订单数"}, inplace=True)
    df_daily_summary = df_daily_summary.merge(df_order_count, on=["文件名", "日期"], how="left")
    df_daily_summary["订单数"] = df_daily_summary["订单数"].fillna(0).astype(int)
    df_daily_summary = df_daily_summary[[
        "文件名", "日期", "订单数", "销售额", "汇率后金额", "P列折后价", "产品成本", "物流成本", "寄样支出", "销量", "寄样数"
    ]]
    df_daily_summary = df_daily_summary.assign(temp_date=pd.to_datetime(df_daily_summary["日期"], errors="coerce")) \
        .sort_values(["文件名", "temp_date"]) \
        .drop(columns=["temp_date"])
    total_row = {"文件名": "全部文件合计", "日期": "汇总"}
    for col in df_daily_summary.columns:
        if col not in ["文件名", "日期"]:
            total_row[col] = df_daily_summary[col].sum()
    df_daily_summary = pd.concat([df_daily_summary, pd.DataFrame([total_row])], ignore_index=True)

    df_daily_product = df_daily_product.assign(temp_date=pd.to_datetime(df_daily_product["日期"], errors="coerce")) \
        .sort_values(["temp_date", "文件名", "产品名称"]) \
        .drop(columns=["temp_date"])
    df_daily_product = df_daily_product[[
        "文件名", "日期", "产品名称", "销售额", "汇率后金额", "P列折后价", "产品成本", "物流成本", "寄样支出", "销量", "寄样数"
    ]]
    df_sku_detail = df_sku_detail.assign(temp_date=pd.to_datetime(df_sku_detail["日期"], errors="coerce")) \
        .sort_values(["temp_date", "文件名", "产品大类", "产品名称"]) \
        .drop(columns=["temp_date"])
    df_sku_detail = df_sku_detail[[
        "文件名", "日期", "产品大类", "产品名称", "订单数", "销售额", "汇率后金额", "P列折后价", "产品成本", "物流成本", "寄样支出", "销量", "寄样数"
    ]]

    period_comparison_frames = build_period_comparison_frames(df_daily_product, daily_order_records)
    df_product_profit_by_period = build_product_profit_by_period(df_daily_product)
    df_product_quantity_by_period = build_product_quantity_by_period(product_quantity_records)
    df_daily_product_quantity_matrix = build_daily_product_quantity_matrix(df_daily_product)

    df_quantity_records = pd.DataFrame(product_quantity_records)
    if not df_quantity_records.empty:
        df_quantity_matrix = df_quantity_records.pivot_table(
            index=["文件名", "产品名称"],
            columns="日期",
            values="销量",
            aggfunc="sum",
            fill_value=0,
        ).astype(int)
        sorted_cols = sorted(df_quantity_matrix.columns, key=lambda x: pd.to_datetime(x, errors="coerce"))
        df_quantity_matrix = df_quantity_matrix[sorted_cols]
        df_quantity_matrix["汇总"] = df_quantity_matrix.sum(axis=1)
        total_row = df_quantity_matrix.sum(axis=0).to_frame().T
        total_row.index = pd.MultiIndex.from_tuples([("全部文件合计", "汇总")], names=df_quantity_matrix.index.names)
        df_quantity_matrix = pd.concat([df_quantity_matrix, total_row]).reset_index()
    else:
        df_quantity_matrix = pd.DataFrame([["无销量数据"]], columns=["提示"])

    df_sample_records = pd.DataFrame(sample_quantity_records)
    if not df_sample_records.empty:
        df_sample_matrix = df_sample_records.pivot_table(
            index=["文件名", "产品名称"],
            columns="日期",
            values="寄样数",
            aggfunc="sum",
            fill_value=0,
        ).astype(int)
        sorted_cols = sorted(df_sample_matrix.columns, key=lambda x: pd.to_datetime(x, errors="coerce"))
        df_sample_matrix = df_sample_matrix[sorted_cols]
        df_sample_matrix["汇总"] = df_sample_matrix.sum(axis=1)
        total_row = df_sample_matrix.sum(axis=0).to_frame().T
        total_row.index = pd.MultiIndex.from_tuples([("全部文件合计", "汇总")], names=df_sample_matrix.index.names)
        df_sample_matrix = pd.concat([df_sample_matrix, total_row]).reset_index()
    else:
        df_sample_matrix = pd.DataFrame([["无样品数据"]], columns=["提示"])

    try:
        os.makedirs(os.path.dirname(output_filename), exist_ok=True)
        with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
            file_summary_sheet = "文件汇总"
            df_file_summary_display = insert_blank_rows_between_files(df_product_profit_by_period)
            df_file_summary_display.to_excel(writer, sheet_name=file_summary_sheet, index=False)
            center_excel_sheet(writer, file_summary_sheet, len(df_file_summary_display) + 1, len(df_file_summary_display.columns))

            detail_sheet = "日产品明细"
            df_daily_product_quantity_matrix.to_excel(writer, sheet_name=detail_sheet, index=False)
            center_excel_sheet(writer, detail_sheet, len(df_daily_product_quantity_matrix) + 1, len(df_daily_product_quantity_matrix.columns))

            sku_detail_sheet = "SKU明细"
            df_sku_detail_display = insert_blank_rows_between_files(df_sku_detail)
            df_sku_detail_display.to_excel(writer, sheet_name=sku_detail_sheet, index=False)
            center_excel_sheet(writer, sku_detail_sheet, len(df_sku_detail_display) + 1, len(df_sku_detail_display.columns))

            quantity_sheet = "产品销量矩阵"
            df_product_quantity_by_period.to_excel(writer, sheet_name=quantity_sheet, index=True)
            quantity_startrow = len(df_product_quantity_by_period) + 3
            df_quantity_matrix_display = insert_blank_rows_between_files(df_quantity_matrix)
            df_quantity_matrix_display.to_excel(writer, sheet_name=quantity_sheet, startrow=quantity_startrow, index=False)
            quantity_rows = quantity_startrow + len(df_quantity_matrix_display) + 1
            quantity_cols = max(len(df_product_quantity_by_period.columns) + 1, len(df_quantity_matrix_display.columns))
            center_excel_sheet(writer, quantity_sheet, quantity_rows, quantity_cols)

            df_sample_matrix_display = insert_blank_rows_between_files(df_sample_matrix)
            df_sample_matrix_display.to_excel(writer, sheet_name="样品统计", index=False)
            center_excel_sheet(writer, "样品统计", len(df_sample_matrix_display) + 1, len(df_sample_matrix_display.columns))

            if period_comparison_frames:
                startrow = 0
                sheet_name = "周期对比"
                max_cols = 0
                for title, frame in period_comparison_frames:
                    frame.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
                    max_cols = max(max_cols, len(frame.columns))
                    startrow += len(frame) + 3
                center_excel_sheet(writer, sheet_name, startrow, max_cols)
            else:
                period_hint = pd.DataFrame([{"提示": "周期对比需要至少 2 个有效 CSV 文件"}])
                period_hint.to_excel(writer, sheet_name="周期对比", index=False)
                center_excel_sheet(writer, "周期对比", len(period_hint) + 1, len(period_hint.columns))
        print(f"✅ [{display_name}] 处理完成，报表已保存至 {output_filename}")

        if unmatched_skus:
            print(f"💡 [{display_name}] 提醒：以下 SKU ID 在配置表中未找到匹配项:")
            for sku in sorted(list(unmatched_skus)):
                print(f"   - {sku}")
    except Exception as e:
        print(f"❌ [{display_name}] 保存失败: {e}")


def run_vietnam_daily(stores=None):
    """运行越南日报。stores 为空时使用脚本内置 VIETNAM_STORES。"""
    print("=" * 60)
    print("TikTok 越南周报生成工具")
    print("=" * 60)
    print(f"✅ 使用固定越南汇率: {VIETNAM_EXCHANGE_RATE}")

    for store in (stores or VIETNAM_STORES):
        if not store.get("enabled", True):
            continue

        sheet_name = store["sheet_name"]
        display_name = f"{store['country_name']}_{store['store_name']}"
        print(f"\n📋 处理组合: {display_name} (汇率: {VIETNAM_EXCHANGE_RATE})")

        try:
            config_df = get_config_dataframe(sheet_name)
            print(f"   ✅ 配置表加载成功，共 {len(config_df)} 条记录")
            run_report(store, config_df, VIETNAM_EXCHANGE_RATE)
        except Exception as e:
            print(f"   ❌ 处理失败: {e}")
            if "没有 SKU 数据" in str(e) or "无法获取配置数据" in str(e):
                print_order_sku_preview(store)
            continue

    print("\n" + "=" * 60)
    print("✨ 越南站所有店铺处理完毕。")
    print("=" * 60)


# ============================================================
# 4. 主程序
# ============================================================
if __name__ == "__main__":
    run_vietnam_daily()
