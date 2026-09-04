# -*- coding: utf-8 -*-
"""
TikTok 墨西哥直邮按日汇总订单成本报表脚本
- 仅处理墨西哥直邮店
- 从飞书表格读取 SKU 映射、产品成本和计费重量
- 物流按订单计费重量汇总后，按重量占比分摊到 SKU / 产品
- 使用固定墨西哥运费价卡 + `.env` 中配置的 MXN->RMB 汇率
"""

import glob
import math
import os
import re
from datetime import datetime, timedelta
from urllib.parse import quote

import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def find_project_root(start_dir):
    env_root = os.getenv("TIKTOK_REPORT_ROOT")
    if env_root:
        root = os.path.abspath(env_root)
        for name in ("config", "data", "result"):
            os.makedirs(os.path.join(root, name), exist_ok=True)
        return root

    current = os.path.abspath(start_dir)
    while True:
        if all(os.path.isdir(os.path.join(current, name)) for name in ("config", "data", "result")):
            return current
        parent = os.path.dirname(current)
        if parent == current:
            raise FileNotFoundError("无法定位 sum_daily_order 根目录，请确认 config/data/result 目录仍在同一层级。")
        current = parent


CURRENT_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = find_project_root(CURRENT_SCRIPT_DIR)


def read_runtime_version():
    version_file = os.getenv("TIKTOK_REPORT_VERSION")
    if version_file:
        return str(version_file).strip().lstrip("vV") or "unversioned"
    for base_dir in (PROJECT_ROOT, CURRENT_SCRIPT_DIR):
        candidate = os.path.join(base_dir, "VERSION")
        if os.path.exists(candidate):
            version_text = open(candidate, encoding="utf-8").read().strip()
            if version_text:
                return version_text.lstrip("vV") or "unversioned"
    return "unversioned"


APP_VERSION = read_runtime_version()

for env_path in [
    os.path.join(PROJECT_ROOT, "config", ".env"),
    os.path.join(CURRENT_SCRIPT_DIR, ".env"),
]:
    if os.path.exists(env_path):
        load_dotenv(env_path)
load_dotenv()


FEISHU_APP_ID = os.getenv("FEISHU_APP_ID")
FEISHU_APP_SECRET = os.getenv("FEISHU_APP_SECRET")
FEISHU_SHEET_TOKEN = os.getenv("FEISHU_SHEET_TOKEN_AMERICAS") or os.getenv("FEISHU_SHEET_TOKEN_MX") or os.getenv("FEISHU_SHEET_TOKEN") or ""
FEISHU_RANGE_SKU = "A:Z"
FEISHU_REQUIRED_SCOPE_HINT = "请在飞书开放平台给应用开通 sheets:spreadsheet:readonly 或 sheets:spreadsheet:read 权限，并重新发布/生效。"

MEXICO_EXCHANGE_RATE = os.getenv("MEXICO_EXCHANGE_RATE")
MX_PER_ORDER_ITEM_FEE_MXN = 6.0
MX_IVA_BASE_RATE = 0.80
MX_IVA_RATE = 0.335
MX_LOCAL_STORE_KEYS = {"local"}
MX_DIRECT_STORE_KEYS = {"direct_old", "direct_new"}

SKU_ID_COLUMN = "SKU ID"
PRODUCT_NAME_COLUMN = "中文简称"
PRODUCT_CATEGORY_COLUMN = "产品大类"
LOGISTICS_COST_COLUMN = "物流成本(元)"
HEAD_LOGISTICS_COST_COLUMN = "头程物流成本(元)"
TAIL_LOGISTICS_COST_COLUMN = "尾程物流成本(元)"
LOCAL_LOGISTICS_UNIT_COLUMN = "_本土店单件物流成本"
SAMPLE_COST_COLUMN = "寄样成本(元)"
LOGISTICS_CAPACITY_COLUMN = "每单物流承载数量"
IMPORT_IVA_COLUMN = "进口IVA"
ACTUAL_WEIGHT_KG_COLUMN = "实重kg"
WEIGHT_KG_COLUMN = "计费重kg"
LENGTH_CM_COLUMN = "长cm"
WIDTH_CM_COLUMN = "宽cm"
HEIGHT_CM_COLUMN = "高cm"
ITEM_QUANTITY_COLUMN = "_ItemQuantity"
MX_VOLUME_WEIGHT_TRIGGER_RATIO = 1.5

CACHE_DIR = os.path.join(PROJECT_ROOT, "config", "cache", APP_VERSION)
os.makedirs(CACHE_DIR, exist_ok=True)
CACHE_EXPIRY_HOURS = 24
USE_CONFIG_CACHE_FIRST = os.getenv("USE_CONFIG_CACHE_FIRST", "").lower() in {"1", "true", "yes", "y"}
SHEET_ID_CACHE = {}

MEXICO_STORES = [
    {
        "enabled": True,
        "country_code": "MX",
        "country_name": "墨西哥",
        "store_key": "local",
        "store_name": "本土店",
        "store_dir": "local",
        "sheet_name": "墨西哥_本土店",
    },
    {
        "enabled": True,
        "country_code": "MX",
        "country_name": "墨西哥",
        "store_key": "direct_old",
        "store_name": "直邮老店",
        "store_dir": "direct_old",
        "sheet_name": "墨西哥_直邮店",
    },
    {
        "enabled": True,
        "country_code": "MX",
        "country_name": "墨西哥",
        "store_key": "direct_new",
        "store_name": "直邮新店",
        "store_dir": "direct_new",
        "sheet_name": "墨西哥_直邮店",
    }
]


def get_mx_exchange_rate():
    if MEXICO_EXCHANGE_RATE is None or str(MEXICO_EXCHANGE_RATE).strip() == "":
        raise EnvironmentError("未配置 MEXICO_EXCHANGE_RATE，无法把墨西哥运费换算成人民币。")
    try:
        return float(MEXICO_EXCHANGE_RATE)
    except Exception as exc:
        raise EnvironmentError(f"MEXICO_EXCHANGE_RATE 不是有效数字: {MEXICO_EXCHANGE_RATE}") from exc


def normalize_feishu_sheet_token(value):
    if not value:
        return ""
    token = str(value).strip().strip("'").strip('"').strip()
    for sep in ("?", "#", "&"):
        if sep in token:
            token = token.split(sep, 1)[0]
    return token


def get_feishu_token(app_id, app_secret):
    url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
    payload = {"app_id": app_id, "app_secret": app_secret}
    try:
        resp = requests.post(url, json=payload, timeout=10)
        if resp.status_code != 200:
            raise Exception(f"飞书Token请求失败，HTTP {resp.status_code}")
        data = resp.json()
        if data.get("code") != 0:
            raise Exception(f"飞书Token错误: {data}")
        return data["tenant_access_token"]
    except Exception as e:
        raise Exception(f"获取飞书 token 失败: {e}")


def format_feishu_error(data):
    if not isinstance(data, dict):
        return str(data)

    code = data.get("code")
    msg = data.get("msg")
    error = data.get("error") or {}
    message = error.get("message")
    log_id = error.get("log_id")

    parts = [f"code={code}", f"msg={msg}"]
    if message:
        parts.append(f"message={message}")
    if log_id:
        parts.append(f"log_id={log_id}")
    if code == 99991672:
        parts.append(FEISHU_REQUIRED_SCOPE_HINT)
    return "；".join(parts)


def get_feishu_json(url, headers, action):
    resp = requests.get(url, headers=headers, timeout=10)
    try:
        data = resp.json()
    except Exception:
        data = {"code": resp.status_code, "msg": resp.text[:500]}

    if resp.status_code != 200 or data.get("code") != 0:
        raise Exception(f"{action}失败，HTTP {resp.status_code}，{format_feishu_error(data)}")
    return data


def get_feishu_sheet_id(token, sheet_token, sheet_name):
    cache_key = (sheet_token, sheet_name)
    if cache_key in SHEET_ID_CACHE:
        return SHEET_ID_CACHE[cache_key]

    headers = {"Authorization": f"Bearer {token}"}
    url = f"https://open.feishu.cn/open-apis/sheets/v3/spreadsheets/{sheet_token}/sheets/query"
    data = get_feishu_json(url, headers, "查询飞书工作表列表")
    sheets = data.get("data", {}).get("sheets", [])

    available = []
    for item in sheets:
        props = item.get("sheet") or item.get("properties") or item
        title = props.get("title") or props.get("sheet_name")
        sheet_id = props.get("sheet_id") or props.get("sheetId")
        if title and sheet_id:
            SHEET_ID_CACHE[(sheet_token, title)] = sheet_id
            SHEET_ID_CACHE[(sheet_token, sheet_id)] = sheet_id
            available.append(f"{title}({sheet_id})")
            if title == sheet_name or sheet_id == sheet_name:
                return sheet_id

    raise ValueError(f"未找到飞书 Sheet '{sheet_name}'。当前可用 Sheet: {', '.join(available) or '无'}")


def read_feishu_sheet(token, sheet_token, sheet_name, range_str):
    headers = {"Authorization": f"Bearer {token}"}
    sheet_id = get_feishu_sheet_id(token, sheet_token, sheet_name)
    range_expr = quote(f"{sheet_id}!{range_str}", safe="")
    url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{sheet_token}/values/{range_expr}"
    try:
        data = get_feishu_json(url, headers, "读取飞书表格")
        values = data.get("data", {}).get("valueRange", {}).get("values", [])
        if not values:
            raise ValueError(f"飞书表格 Sheet '{sheet_name}' 返回空数据")
        return pd.DataFrame(values[1:], columns=values[0])
    except Exception as e:
        raise Exception(f"读取飞书表格 Sheet '{sheet_name}' 失败: {e}")


def clean_sku_id_column(df):
    if SKU_ID_COLUMN not in df.columns:
        return df
    df = df.dropna(subset=[SKU_ID_COLUMN]).copy()
    df[SKU_ID_COLUMN] = df[SKU_ID_COLUMN].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    df = df[df[SKU_ID_COLUMN] != ""]
    return df


def clean_config_dataframe(df):
    df = clean_sku_id_column(df)
    if "SKU中文简称" in df.columns and PRODUCT_NAME_COLUMN not in df.columns:
        df = df.rename(columns={"SKU中文简称": PRODUCT_NAME_COLUMN})
    if ACTUAL_WEIGHT_KG_COLUMN not in df.columns:
        if "计费重量kg" in df.columns:
            df = df.rename(columns={"计费重量kg": ACTUAL_WEIGHT_KG_COLUMN})
        elif "计费重量g" in df.columns:
            df = df.rename(columns={"计费重量g": ACTUAL_WEIGHT_KG_COLUMN})

    if WEIGHT_KG_COLUMN not in df.columns:
        if "计费重量kg" in df.columns:
            df = df.rename(columns={"计费重量kg": WEIGHT_KG_COLUMN})
        elif "计费重量g" in df.columns:
            df = df.rename(columns={"计费重量g": WEIGHT_KG_COLUMN})

    if PRODUCT_CATEGORY_COLUMN not in df.columns:
        df[PRODUCT_CATEGORY_COLUMN] = df.get(PRODUCT_NAME_COLUMN, "")
    if PRODUCT_NAME_COLUMN in df.columns:
        df[PRODUCT_NAME_COLUMN] = df[PRODUCT_NAME_COLUMN].astype(str).str.strip()
    if PRODUCT_CATEGORY_COLUMN in df.columns:
        df[PRODUCT_CATEGORY_COLUMN] = df[PRODUCT_CATEGORY_COLUMN].astype(str).str.strip()
        df[PRODUCT_CATEGORY_COLUMN] = df[PRODUCT_CATEGORY_COLUMN].replace({"": pd.NA, "nan": pd.NA})
        if PRODUCT_NAME_COLUMN in df.columns:
            df[PRODUCT_CATEGORY_COLUMN] = df[PRODUCT_CATEGORY_COLUMN].fillna(df[PRODUCT_NAME_COLUMN])

    if ACTUAL_WEIGHT_KG_COLUMN not in df.columns:
        df[ACTUAL_WEIGHT_KG_COLUMN] = 0.0
    if WEIGHT_KG_COLUMN not in df.columns:
        df[WEIGHT_KG_COLUMN] = 0.0
    df[ACTUAL_WEIGHT_KG_COLUMN] = pd.to_numeric(df[ACTUAL_WEIGHT_KG_COLUMN], errors="coerce").fillna(0.0)
    df[WEIGHT_KG_COLUMN] = pd.to_numeric(df[WEIGHT_KG_COLUMN], errors="coerce").fillna(0.0)
    df.loc[df[ACTUAL_WEIGHT_KG_COLUMN] <= 0, ACTUAL_WEIGHT_KG_COLUMN] = 0.0
    df.loc[df[WEIGHT_KG_COLUMN] <= 0, WEIGHT_KG_COLUMN] = 0.0

    for col in [LENGTH_CM_COLUMN, WIDTH_CM_COLUMN, HEIGHT_CM_COLUMN]:
        if col not in df.columns:
            df[col] = 0.0
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
        df.loc[df[col] < 0, col] = 0.0

    if "产品成本(元)" not in df.columns:
        df["产品成本(元)"] = 0.0
    df["产品成本(元)"] = pd.to_numeric(df["产品成本(元)"], errors="coerce").fillna(0.0)
    for col in [LOGISTICS_COST_COLUMN, HEAD_LOGISTICS_COST_COLUMN, TAIL_LOGISTICS_COST_COLUMN, SAMPLE_COST_COLUMN, LOGISTICS_CAPACITY_COLUMN, IMPORT_IVA_COLUMN]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    return df


def get_cache_file_path(sheet_name):
    safe_name = sheet_name.replace(" ", "_").replace("/", "_")
    return os.path.join(CACHE_DIR, f"config_{safe_name}.csv")


def get_config_dataframe(sheet_name, force_refresh=False):
    cache_file = get_cache_file_path(sheet_name)

    try:
        if not FEISHU_APP_ID or not FEISHU_APP_SECRET or not FEISHU_SHEET_TOKEN:
            raise EnvironmentError("未配置 FEISHU_APP_ID / FEISHU_APP_SECRET / FEISHU_SHEET_TOKEN_AMERICAS，无法从飞书刷新配置。")
        token = get_feishu_token(FEISHU_APP_ID, FEISHU_APP_SECRET)
        df = read_feishu_sheet(token, FEISHU_SHEET_TOKEN, sheet_name, FEISHU_RANGE_SKU)
        df = df.dropna(how="all")
        df = clean_config_dataframe(df)
        df.to_csv(cache_file, index=False, encoding="utf-8-sig")
        print(f"   ✅ 成功从飞书拉取配置并缓存（Sheet: {sheet_name}）")
        return df
    except Exception as e:
        print(f"   ❌ 飞书拉取失败（Sheet: {sheet_name}）: {e}")
        if os.path.exists(cache_file):
            try:
                df = pd.read_csv(cache_file, dtype=str, encoding="utf-8-sig")
                df = clean_config_dataframe(df)
                print(f"   ⚠️ 飞书不可用，使用过期缓存（Sheet: {sheet_name}）")
                return df
            except Exception:
                pass
        raise Exception(f"无法获取配置数据（Sheet: {sheet_name}）")


def parse_amount(value):
    if pd.isna(value):
        return 0.0
    s = str(value).replace(" ", "").replace("\t", "").replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", s)
    if not match:
        return 0.0
    try:
        return float(match.group(0))
    except Exception:
        return 0.0


def clean_sku_str(sku_series):
    return sku_series.astype(str).str.replace(r"\.0$", "", regex=True).str.strip()


def drop_conflicting_config_columns(df, config_cols):
    conflict_cols = [col for col in config_cols if col in df.columns and col != SKU_ID_COLUMN]
    if conflict_cols:
        df = df.drop(columns=conflict_cols)
    return df


def parse_order_date(value):
    if pd.isna(value):
        return pd.NA
    s = str(value).strip()
    if not s:
        return pd.NA
    try:
        dt = pd.to_datetime(s, format="%m/%d/%Y %I:%M:%S %p", errors="raise")
        return dt.strftime("%Y-%m-%d")
    except Exception:
        try:
            dt = pd.to_datetime(s, errors="coerce")
            if pd.isna(dt):
                return pd.NA
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return pd.NA


def get_item_quantity(df):
    if "Quantity" not in df.columns:
        return pd.Series(1, index=df.index, dtype="float64")
    qty = pd.to_numeric(df["Quantity"], errors="coerce").fillna(1)
    qty.loc[qty <= 0] = 1
    return qty


def get_order_key(df):
    if "Order ID" in df.columns:
        order_key = df["Order ID"].astype(str).str.strip()
        missing_order = order_key.isna() | order_key.eq("") | order_key.eq("nan")
        return order_key.mask(missing_order, "ROW_" + df.index.astype(str))
    return "ROW_" + df.index.astype(str)


def validate_config_columns(config_df, sheet_name, store_key=None):
    required_cols = [SKU_ID_COLUMN, PRODUCT_NAME_COLUMN, "产品成本(元)", PRODUCT_CATEGORY_COLUMN]
    if store_key in MX_LOCAL_STORE_KEYS:
        required_cols.extend([HEAD_LOGISTICS_COST_COLUMN, TAIL_LOGISTICS_COST_COLUMN])
    else:
        required_cols.extend([WEIGHT_KG_COLUMN])
    missing = [col for col in required_cols if col not in config_df.columns]
    if missing:
        raise ValueError(f"配置表 {sheet_name} 缺少列: {', '.join(missing)}")
    if config_df.empty:
        raise ValueError(f"配置表 {sheet_name} 没有 SKU 数据，请先补充 SKU ID、SKU中文简称和成本信息")


def get_config_merge_columns(config_df, store_key, include_sample_cost=True):
    cols = [SKU_ID_COLUMN, PRODUCT_CATEGORY_COLUMN, PRODUCT_NAME_COLUMN, "产品成本(元)"]
    if IMPORT_IVA_COLUMN in config_df.columns:
        cols.append(IMPORT_IVA_COLUMN)
    if store_key in MX_LOCAL_STORE_KEYS:
        for col in [HEAD_LOGISTICS_COST_COLUMN, TAIL_LOGISTICS_COST_COLUMN, LOGISTICS_CAPACITY_COLUMN]:
            if col in config_df.columns:
                cols.append(col)
    else:
        for col in [ACTUAL_WEIGHT_KG_COLUMN, WEIGHT_KG_COLUMN, LENGTH_CM_COLUMN, WIDTH_CM_COLUMN, HEIGHT_CM_COLUMN]:
            if col in config_df.columns:
                cols.append(col)
    return cols


def safe_change_rate(current, previous):
    if previous == 0:
        return ""
    return round((current - previous) / previous, 4)


def format_percent(value):
    if value == "" or pd.isna(value):
        return ""
    return f"{value:.2%}"


def build_order_logistics_allocation(df_normal, logistics_col):
    result_cols = ["日期", "Product Category", "Mapped Name", "物流成本"]
    if df_normal.empty:
        return pd.DataFrame(columns=result_cols)

    df_calc = df_normal.copy()
    df_calc["_OrderKey"] = get_order_key(df_calc)
    qty_source = df_calc[ITEM_QUANTITY_COLUMN] if ITEM_QUANTITY_COLUMN in df_calc.columns else pd.Series(1, index=df_calc.index)
    df_calc[ITEM_QUANTITY_COLUMN] = pd.to_numeric(qty_source, errors="coerce").fillna(1)
    df_calc.loc[df_calc[ITEM_QUANTITY_COLUMN] <= 0, ITEM_QUANTITY_COLUMN] = 1
    logistics_source = df_calc[logistics_col] if logistics_col in df_calc.columns else pd.Series(0, index=df_calc.index)
    df_calc["_LineLogistics"] = pd.to_numeric(logistics_source, errors="coerce").fillna(0) * df_calc[ITEM_QUANTITY_COLUMN]
    capacity_source = df_calc[LOGISTICS_CAPACITY_COLUMN] if LOGISTICS_CAPACITY_COLUMN in df_calc.columns else pd.Series(1, index=df_calc.index)
    df_calc[LOGISTICS_CAPACITY_COLUMN] = pd.to_numeric(capacity_source, errors="coerce").fillna(1)
    df_calc.loc[df_calc[LOGISTICS_CAPACITY_COLUMN] <= 0, LOGISTICS_CAPACITY_COLUMN] = 1

    category_order = df_calc.groupby(["_OrderKey", "日期", "Product Category"]).agg(
        category_qty=(ITEM_QUANTITY_COLUMN, "sum"),
        original_logistics=("_LineLogistics", "sum"),
        unit_logistics=(logistics_col, "max"),
        capacity=(LOGISTICS_CAPACITY_COLUMN, "max"),
    ).reset_index()
    category_order["_capacity_logistics"] = category_order.apply(
        lambda row: math.ceil(row["category_qty"] / row["capacity"]) * row["unit_logistics"],
        axis=1,
    )
    category_order["candidate_logistics"] = category_order[["original_logistics", "_capacity_logistics"]].min(axis=1)
    order_max = category_order.groupby("_OrderKey")["candidate_logistics"].transform("max")
    order_sum = category_order.groupby("_OrderKey")["candidate_logistics"].transform("sum")
    category_order["category_logistics"] = category_order.apply(
        lambda row: 0 if order_sum.loc[row.name] == 0 else order_max.loc[row.name] * row["candidate_logistics"] / order_sum.loc[row.name],
        axis=1,
    )

    sku_order = df_calc.groupby(["_OrderKey", "日期", "Product Category", "Mapped Name"]).agg(
        sku_qty=(ITEM_QUANTITY_COLUMN, "sum"),
        sku_original_logistics=("_LineLogistics", "sum"),
    ).reset_index()
    sku_order = sku_order.merge(
        category_order[["_OrderKey", "日期", "Product Category", "category_qty", "original_logistics", "category_logistics"]],
        on=["_OrderKey", "日期", "Product Category"],
        how="left",
    )
    sku_order["物流成本"] = sku_order.apply(
        lambda row: (
            0
            if row["category_qty"] == 0
            else row["category_logistics"] * (
                row["sku_original_logistics"] / row["original_logistics"]
                if row["original_logistics"] != 0
                else row["sku_qty"] / row["category_qty"]
            )
        ),
        axis=1,
    )
    return sku_order.groupby(["日期", "Product Category", "Mapped Name"])["物流成本"].sum().reset_index()


def build_mx_local_unit_logistics(df):
    df_calc = df.copy()
    head = pd.to_numeric(df_calc[HEAD_LOGISTICS_COST_COLUMN] if HEAD_LOGISTICS_COST_COLUMN in df_calc.columns else pd.Series(0, index=df_calc.index), errors="coerce").fillna(0)
    tail = pd.to_numeric(df_calc[TAIL_LOGISTICS_COST_COLUMN] if TAIL_LOGISTICS_COST_COLUMN in df_calc.columns else pd.Series(0, index=df_calc.index), errors="coerce").fillna(0)
    df_calc[LOCAL_LOGISTICS_UNIT_COLUMN] = head + tail
    return df_calc


def get_mx_freight_cost_mxn(billable_weight_kg):
    rate_cards = [
        (0.1, 69),
        (0.2, 93),
        (0.3, 114),
        (0.4, 134),
        (0.5, 155),
        (0.6, 176),
        (0.7, 197),
        (0.8, 218),
        (0.9, 239),
        (1.0, 259),
        (1.5, 350),
        (2.0, 454),
        (3.0, 616),
        (4.0, 824),
        (5.0, 1033),
        (6.0, 1394),
        (7.0, 1603),
        (8.0, 1811),
        (9.0, 1825),
        (10.0, 2059),
        (11.0, 2293),
        (12.0, 2527),
        (13.0, 2762),
        (14.0, 2996),
    ]
    weight = max(float(billable_weight_kg or 0), 0.0)
    if weight <= 0:
        return 0
    for upper_bound, price in rate_cards:
        if weight <= upper_bound:
            return price
    return 3230


def calc_mexico_billable_weight_kg(actual_weight_kg, length_cm, width_cm, height_cm, quantity=1, billed_weight_kg=None):
    qty = max(float(quantity or 0), 0.0)
    actual_weight = max(float(actual_weight_kg or 0), 0.0)
    length = max(float(length_cm or 0), 0.0)
    width = max(float(width_cm or 0), 0.0)
    height = max(float(height_cm or 0), 0.0)
    billed_weight = max(float(billed_weight_kg or 0), 0.0)

    if actual_weight <= 0 or length <= 0 or width <= 0 or height <= 0:
        return 0.0

    actual_weight = actual_weight * qty
    volume_weight = (length * width * height / 8000.0) * qty
    if billed_weight > 0:
        return billed_weight * qty
    if volume_weight > actual_weight * MX_VOLUME_WEIGHT_TRIGGER_RATIO:
        return volume_weight
    return actual_weight


def calc_mexico_iva_amount(import_iva_value, fallback_base_amount, quantity=1):
    iva_value = parse_amount(import_iva_value)
    qty = max(float(quantity or 0), 0.0)
    if iva_value != 0:
        return round(abs(iva_value) * max(qty, 1.0), 2)
    base_amount = max(float(fallback_base_amount or 0), 0.0)
    return round(base_amount * MX_IVA_BASE_RATE * MX_IVA_RATE, 2)


def build_mexico_direct_logistics_allocation(df_normal, exchange_rate):
    result_cols = ["日期", "Product Category", "Mapped Name", "物流成本"]
    if df_normal.empty:
        return pd.DataFrame(columns=result_cols)

    df_calc = df_normal.copy()
    if "Order ID" in df_calc.columns:
        order_key = df_calc["Order ID"].astype(str).str.strip()
        missing_order = order_key.isna() | order_key.eq("") | order_key.eq("nan")
        order_key = order_key.mask(missing_order, "ROW_" + df_calc.index.astype(str))
    else:
        order_key = "ROW_" + df_calc.index.astype(str)
    df_calc["_OrderKey"] = order_key
    if ITEM_QUANTITY_COLUMN in df_calc.columns:
        df_calc[ITEM_QUANTITY_COLUMN] = pd.to_numeric(df_calc[ITEM_QUANTITY_COLUMN], errors="coerce").fillna(1)
    else:
        df_calc[ITEM_QUANTITY_COLUMN] = 1
    df_calc.loc[df_calc[ITEM_QUANTITY_COLUMN] <= 0, ITEM_QUANTITY_COLUMN] = 1
    if WEIGHT_KG_COLUMN in df_calc.columns:
        df_calc[WEIGHT_KG_COLUMN] = pd.to_numeric(df_calc[WEIGHT_KG_COLUMN], errors="coerce").fillna(0)
    else:
        df_calc[WEIGHT_KG_COLUMN] = 0
    df_calc.loc[df_calc[WEIGHT_KG_COLUMN] < 0, WEIGHT_KG_COLUMN] = 0
    for col in [LENGTH_CM_COLUMN, WIDTH_CM_COLUMN, HEIGHT_CM_COLUMN]:
        if col in df_calc.columns:
            df_calc[col] = pd.to_numeric(df_calc[col], errors="coerce").fillna(0)
        else:
            df_calc[col] = 0
        df_calc.loc[df_calc[col] < 0, col] = 0
    df_calc["_LineWeightKg"] = df_calc.apply(
        lambda row: calc_mexico_billable_weight_kg(
            row[ACTUAL_WEIGHT_KG_COLUMN],
            row[LENGTH_CM_COLUMN],
            row[WIDTH_CM_COLUMN],
            row[HEIGHT_CM_COLUMN],
            row[ITEM_QUANTITY_COLUMN],
            row[WEIGHT_KG_COLUMN],
        ),
        axis=1,
    )

    order_summary = df_calc.groupby(["_OrderKey", "日期"]).agg(
        order_weight_kg=("_LineWeightKg", "sum"),
        order_qty=(ITEM_QUANTITY_COLUMN, "sum"),
    ).reset_index()
    order_summary["order_logistics_mxn"] = order_summary["order_weight_kg"].map(get_mx_freight_cost_mxn)
    order_summary["order_logistics"] = order_summary["order_logistics_mxn"] * exchange_rate

    df_calc = df_calc.merge(
        order_summary[["_OrderKey", "日期", "order_weight_kg", "order_qty", "order_logistics"]],
        on=["_OrderKey", "日期"],
        how="left",
    )
    df_calc["_WeightShare"] = df_calc.apply(
        lambda row: (
            row["_LineWeightKg"] / row["order_weight_kg"]
            if row["order_weight_kg"] > 0
            else row[ITEM_QUANTITY_COLUMN] / row["order_qty"]
            if row["order_qty"] > 0
            else 0
        ),
        axis=1,
    )
    df_calc["物流成本"] = df_calc["order_logistics"] * df_calc["_WeightShare"]
    return df_calc.groupby(["日期", "Product Category", "Mapped Name"])["物流成本"].sum().reset_index()


def build_period_comparison_frames(df_daily_product, daily_order_records, exchange_rate):
    if df_daily_product.empty or "文件名" not in df_daily_product.columns:
        return []

    period_summary = df_daily_product.groupby("文件名").agg({
        "销量": "sum",
        "寄样数": "sum",
        "销售额": "sum",
        "汇率后金额": "sum",
        "除运费外销售额": "sum",
        "产品成本": "sum",
        "物流成本": "sum",
        "正常订单IVA": "sum",
        "寄样支出": "sum",
        "寄样IVA": "sum",
        "寄样总成本": "sum",
        "日期": ["min", "max"],
    })
    period_summary.columns = [
        "销量", "寄样数", "销售额", "汇率后金额", "除运费外销售额",
        "产品成本", "物流成本", "正常订单IVA", "寄样支出", "寄样IVA", "寄样总成本",
        "开始日期", "结束日期"
    ]

    if daily_order_records:
        df_order_count = (
            pd.DataFrame(daily_order_records)
            .drop_duplicates(subset=["文件名", "日期", "Order ID"])
            .groupby("文件名")["Order ID"]
            .count()
            .rename("订单数")
        )
        period_summary = period_summary.join(df_order_count, how="left")
    else:
        period_summary["订单数"] = period_summary["销量"]

    period_summary["订单数"] = period_summary["订单数"].fillna(0)
    period_summary["每件商品成交费用"] = period_summary["销量"] * MX_PER_ORDER_ITEM_FEE_MXN * exchange_rate
    period_summary["利润"] = (
        period_summary["汇率后金额"]
        - period_summary["产品成本"]
        - period_summary["物流成本"]
        - period_summary["正常订单IVA"]
        - period_summary["寄样总成本"]
        - period_summary["每件商品成交费用"]
    )
    period_summary["利润率"] = period_summary.apply(
        lambda row: 0 if row["汇率后金额"] == 0 else row["利润"] / row["汇率后金额"],
        axis=1,
    )
    period_summary = period_summary.reset_index()
    period_summary["_开始日期排序"] = pd.to_datetime(period_summary["开始日期"], errors="coerce")
    period_summary["_结束日期排序"] = pd.to_datetime(period_summary["结束日期"], errors="coerce")
    period_summary = (
        period_summary
        .sort_values(["_开始日期排序", "_结束日期排序", "文件名"])
        .drop(columns=["_开始日期排序", "_结束日期排序"])
        .set_index("文件名")
    )

    metrics = ["订单数", "销售额", "汇率后金额", "除运费外销售额", "产品成本", "物流成本", "正常订单IVA", "寄样支出", "寄样IVA", "寄样总成本", "每件商品成交费用", "利润", "利润率", "销量", "寄样数"]
    frames = []
    ordered_files = list(period_summary.index)

    summary_rows = []
    for file_name in ordered_files:
        row = {"文件名": file_name}
        for metric in metrics:
            row[metric] = format_percent(period_summary.loc[file_name, metric]) if metric == "利润率" else round(period_summary.loc[file_name, metric], 2)
        summary_rows.append(row)
    frames.append(("周期汇总", pd.DataFrame(summary_rows)))

    for idx in range(1, len(ordered_files)):
        previous_file = ordered_files[idx - 1]
        current_file = ordered_files[idx]
        previous = period_summary.loc[previous_file]
        current = period_summary.loc[current_file]
        change_value_row = {"文件名": f"变化值：{current_file} - {previous_file}"}
        change_rate_row = {"文件名": f"变化率：{current_file} / {previous_file} - 1"}
        for metric in metrics:
            if metric == "利润率":
                change_value_row[metric] = format_percent(current[metric] - previous[metric])
                change_rate_row[metric] = ""
            else:
                previous_value = round(previous[metric], 2)
                current_value = round(current[metric], 2)
                change_value_row[metric] = round(current_value - previous_value, 2)
                change_rate_row[metric] = format_percent(safe_change_rate(current_value, previous_value))
        title = f"{current_file} - {previous_file}"
        frames.append((title, pd.DataFrame([change_value_row, change_rate_row])))
    return frames


def center_excel_sheet(writer, sheet_name, row_count, col_count):
    worksheet = writer.sheets[sheet_name]
    if hasattr(worksheet, "set_column"):
        worksheet.freeze_panes(1, 0)
        if not hasattr(writer, "_center_format"):
            writer._center_format = writer.book.add_format({"align": "center", "valign": "vcenter"})
        center_format = writer._center_format
        worksheet.set_column(0, max(col_count - 1, 0), 14, center_format)
        for row_idx in range(row_count):
            worksheet.set_row(row_idx, None, center_format)
        return

    worksheet.freeze_panes = "A2"
    center_alignment = Alignment(horizontal="center", vertical="center")
    for row in worksheet.iter_rows(min_row=1, max_row=row_count, min_col=1, max_col=col_count):
        for cell in row:
            cell.alignment = center_alignment

    for col_idx in range(1, col_count + 1):
        max_width = 10
        for row_idx in range(1, row_count + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            text = str(value)
            width = sum(2 if ord(char) > 127 else 1 for char in text)
            max_width = max(max_width, width + 2)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_width, 28)

    for row_idx in range(1, row_count + 1):
        worksheet.row_dimensions[row_idx].height = 20


def insert_blank_rows_between_files(df):
    if df.empty or "文件名" not in df.columns:
        return df

    summary_labels = {"全部文件合计", "汇总", "总计"}
    parts = []
    blank_row = {col: "" for col in df.columns}
    for file_name, group in df.groupby("文件名", sort=False, dropna=False):
        if str(file_name).strip() in summary_labels:
            parts.append(group)
            continue
        if parts:
            parts.append(pd.DataFrame([blank_row], columns=df.columns))
        parts.append(group)
    return pd.concat(parts, ignore_index=True)


def build_product_profit_by_period(df_daily_product, exchange_rate):
    if df_daily_product.empty:
        return pd.DataFrame([["无产品汇总数据"]], columns=["提示"])

    summary = df_daily_product.groupby(["文件名", "产品名称"]).agg({
        "销量": "sum",
        "寄样数": "sum",
        "销售额": "sum",
        "汇率后金额": "sum",
        "除运费外销售额": "sum",
        "产品成本": "sum",
        "物流成本": "sum",
        "正常订单IVA": "sum",
        "寄样支出": "sum",
        "寄样IVA": "sum",
        "寄样总成本": "sum",
    }).reset_index()
    summary["每件商品成交费用"] = summary["销量"] * MX_PER_ORDER_ITEM_FEE_MXN * exchange_rate
    summary["总成本"] = (
        summary["产品成本"]
        + summary["物流成本"]
        + summary["正常订单IVA"]
        + summary["寄样总成本"]
        + summary["每件商品成交费用"]
    )
    summary["利润"] = summary["汇率后金额"] - summary["总成本"]
    summary["利润率"] = summary.apply(lambda row: "" if row["汇率后金额"] == 0 else format_percent(row["利润"] / row["汇率后金额"]), axis=1)
    for col in ["销售额", "汇率后金额", "除运费外销售额", "产品成本", "物流成本", "正常订单IVA", "寄样支出", "寄样IVA", "寄样总成本", "每件商品成交费用", "总成本", "利润"]:
        summary[col] = summary[col].round(2)
    return summary[[
        "文件名", "产品名称", "销售额", "汇率后金额", "除运费外销售额",
        "产品成本", "物流成本", "正常订单IVA", "寄样支出", "寄样IVA", "寄样总成本", "每件商品成交费用", "总成本", "利润", "利润率"
    ]]


def build_sku_profit_by_period(df_sku_detail, exchange_rate):
    if df_sku_detail.empty:
        return pd.DataFrame([["无 SKU 汇总数据"]], columns=["提示"])

    summary = df_sku_detail.groupby(["文件名", "产品大类", "产品名称"]).agg({
        "销量": "sum",
        "寄样数": "sum",
        "销售额": "sum",
        "汇率后金额": "sum",
        "除运费外销售额": "sum",
        "产品成本": "sum",
        "物流成本": "sum",
        "正常订单IVA": "sum",
        "寄样支出": "sum",
        "寄样IVA": "sum",
        "寄样总成本": "sum",
    }).reset_index()
    summary["每件商品成交费用"] = summary["销量"] * MX_PER_ORDER_ITEM_FEE_MXN * exchange_rate
    summary["总成本"] = (
        summary["产品成本"]
        + summary["物流成本"]
        + summary["正常订单IVA"]
        + summary["寄样总成本"]
        + summary["每件商品成交费用"]
    )
    summary["利润"] = summary["汇率后金额"] - summary["总成本"]
    summary["利润率"] = summary.apply(lambda row: "" if row["汇率后金额"] == 0 else format_percent(row["利润"] / row["汇率后金额"]), axis=1)
    for col in ["销售额", "汇率后金额", "除运费外销售额", "产品成本", "物流成本", "正常订单IVA", "寄样支出", "寄样IVA", "寄样总成本", "每件商品成交费用", "总成本", "利润"]:
        summary[col] = summary[col].round(2)
    return summary[[
        "文件名", "产品大类", "产品名称", "销售额", "汇率后金额", "除运费外销售额",
        "产品成本", "物流成本", "正常订单IVA", "寄样支出", "寄样IVA", "寄样总成本", "每件商品成交费用", "总成本", "利润", "利润率", "销量", "寄样数"
    ]]


def build_product_quantity_by_period(product_quantity_records):
    df_quantity_records = pd.DataFrame(product_quantity_records)
    if df_quantity_records.empty:
        return pd.DataFrame([["无销量数据"]], columns=["提示"])

    matrix = df_quantity_records.pivot_table(
        index="产品名称",
        columns="文件名",
        values="销量",
        aggfunc="sum",
        fill_value=0,
    ).astype(int)
    matrix["汇总"] = matrix.sum(axis=1)
    matrix.loc["汇总"] = matrix.sum(axis=0)
    return matrix


def build_daily_product_quantity_matrix(df_daily_product):
    if df_daily_product.empty:
        return pd.DataFrame([["无每日产品销量数据"]], columns=["提示"])

    matrix = df_daily_product.pivot_table(
        index="产品名称",
        columns="日期",
        values="销量",
        aggfunc="sum",
        fill_value=0,
    ).astype(int)
    sorted_cols = sorted(matrix.columns, key=lambda x: pd.to_datetime(x, errors="coerce"))
    matrix = matrix[sorted_cols]
    matrix["汇总"] = matrix.sum(axis=1)
    matrix.loc["汇总"] = matrix.sum(axis=0)
    return matrix.reset_index()


def build_daily_sku_quantity_matrix(df_sku_detail):
    if df_sku_detail.empty:
        return pd.DataFrame([["无每日 SKU 销量数据"]], columns=["提示"])

    matrix = df_sku_detail.pivot_table(
        index=["产品大类", "产品名称"],
        columns="日期",
        values="销量",
        aggfunc="sum",
        fill_value=0,
    ).astype(int)
    sorted_cols = sorted(matrix.columns, key=lambda x: pd.to_datetime(x, errors="coerce"))
    matrix = matrix[sorted_cols]
    matrix["汇总"] = matrix.sum(axis=1)
    total_row = matrix.sum(axis=0).to_frame().T
    total_row.index = pd.MultiIndex.from_tuples([("汇总", "汇总")], names=matrix.index.names)
    matrix = pd.concat([matrix, total_row]).reset_index()
    return matrix.rename(columns={"产品名称": "产品SKU"})


def collect_order_sku_preview(store_config, max_rows=80):
    country_code = store_config["country_code"]
    store_dir = store_config["store_dir"]
    folder_path = os.path.normpath(os.path.join(PROJECT_ROOT, "data", f"data_{country_code}", store_dir))
    file_list = glob.glob(os.path.join(folder_path, "*.csv"))

    preview = {}
    for file_path in file_list:
        try:
            df = pd.read_csv(file_path, dtype=str)
            df.columns = df.columns.astype(str).str.strip()
        except Exception:
            continue

        if SKU_ID_COLUMN not in df.columns:
            continue

        for _, row in df.iterrows():
            sku_id = str(row.get(SKU_ID_COLUMN, "")).replace(".0", "").strip()
            if not sku_id or sku_id.lower() == "nan" or sku_id in preview:
                continue

            preview[sku_id] = {
                "Seller SKU": str(row.get("Seller SKU", "")).strip(),
                "Product Name": str(row.get("Product Name", "")).strip(),
            }
            if len(preview) >= max_rows:
                return preview
    return preview


def print_order_sku_preview(store_config):
    preview = collect_order_sku_preview(store_config)
    if not preview:
        return

    display_name = f"{store_config['country_name']}_{store_config['store_name']}"
    print(f"   💡 [{display_name}] 订单中出现的 SKU ID，可用于补充配置表:")
    for sku_id, info in preview.items():
        seller_sku = info["Seller SKU"]
        product_name = info["Product Name"]
        if len(product_name) > 60:
            product_name = product_name[:60] + "..."
        print(f"      - {sku_id} | {seller_sku} | {product_name}")


def run_report(store_config, config_df, exchange_rate):
    country_code = store_config["country_code"]
    country_name = store_config["country_name"]
    store_key = store_config["store_key"]
    store_name = store_config["store_name"]
    store_dir = store_config["store_dir"]
    display_name = f"{country_name}_{store_name}"

    validate_config_columns(config_df, store_config["sheet_name"], store_key)
    is_local_store = store_key in MX_LOCAL_STORE_KEYS

    folder_path = os.path.normpath(os.path.join(PROJECT_ROOT, "data", f"data_{country_code}", store_dir))
    output_filename = os.path.normpath(os.path.join(PROJECT_ROOT, f"result/Daily_Performance_Report_{country_code.upper()}_{store_key}.xlsx"))
    file_list = glob.glob(os.path.join(folder_path, "*.csv"))
    if not file_list:
        print(f"⚠️  [{display_name}] 路径下未找到文件: {folder_path}")
        return

    print(f"🚀 开始处理 [{display_name}]，共 {len(file_list)} 个文件")

    status_col = "Order Status"
    quantity_col = "Normal or Pre-order"
    n_col = "SKU Platform Discount"
    p_col = "SKU Subtotal After Discount"
    r_col = "Original Shipping Fee"
    p_display_col = "除运费外销售额"

    daily_product_detail_list = []
    sku_order_detail_list = []
    product_quantity_records = []
    sample_quantity_records = []
    daily_order_records = []
    unmatched_skus = set()
    missing_cost_skus = set()
    missing_weight_skus = set()

    config_merge_cols = get_config_merge_columns(config_df, store_key, include_sample_cost=True)

    for file_path in file_list:
        file_name = os.path.basename(file_path)
        try:
            df = pd.read_csv(file_path)
        except Exception as e:
            print(f"读取失败 {file_name}: {e}")
            continue

        required_order_cols = [status_col, quantity_col, SKU_ID_COLUMN, n_col, p_col, r_col]
        missing_order_cols = [col for col in required_order_cols if col not in df.columns]
        if missing_order_cols:
            print(f"⚠️ 文件 {file_name} 缺少必要列: {', '.join(missing_order_cols)}，跳过")
            continue

        cancel_keywords = ["Canceled", "Unpaid", "キャンセル済み", "未払い", "已取消", "未支付", "部分发货后取消"]
        status_series = df[status_col].astype(str).str.strip()
        quantity_series = df[quantity_col].astype(str).str.strip()
        normal_mask = ~status_series.isin(cancel_keywords) & quantity_series.eq("Normal")

        df_normal = df[normal_mask].copy()
        if "Paid Time" not in df_normal.columns:
            print(f"⚠️ 文件 {file_name} 中未找到 Paid Time 列，将跳过此文件")
            continue
        df_normal["日期"] = df_normal["Paid Time"].map(parse_order_date)
        df_normal = df_normal.dropna(subset=["日期"]).copy()

        for col in [n_col, p_col, r_col]:
            df_normal[col] = df_normal[col].map(parse_amount)
        df_normal[p_display_col] = df_normal[n_col] + df_normal[p_col]
        df_normal["SKU_ID"] = clean_sku_str(df_normal[SKU_ID_COLUMN])
        if ITEM_QUANTITY_COLUMN in df_normal.columns:
            df_normal[ITEM_QUANTITY_COLUMN] = pd.to_numeric(df_normal[ITEM_QUANTITY_COLUMN], errors="coerce").fillna(1)
        else:
            df_normal[ITEM_QUANTITY_COLUMN] = 1
        df_normal.loc[df_normal[ITEM_QUANTITY_COLUMN] <= 0, ITEM_QUANTITY_COLUMN] = 1

        df_normal = drop_conflicting_config_columns(df_normal, config_merge_cols)
        df_normal = df_normal.merge(config_df[config_merge_cols], left_on="SKU_ID", right_on=SKU_ID_COLUMN, how="left")
        missing = df_normal[df_normal[PRODUCT_NAME_COLUMN].isna()]
        if not missing.empty:
            for sku in missing["SKU_ID"].unique():
                unmatched_skus.add(sku)
        cost_missing_mask = df_normal["产品成本(元)"].fillna(0) <= 0
        if cost_missing_mask.any():
            for sku in df_normal.loc[cost_missing_mask, "SKU_ID"].unique():
                missing_cost_skus.add(sku)
        if not is_local_store:
            weight_missing_mask = (
                (df_normal[ACTUAL_WEIGHT_KG_COLUMN].fillna(0) <= 0)
                | (df_normal[LENGTH_CM_COLUMN].fillna(0) <= 0)
                | (df_normal[WIDTH_CM_COLUMN].fillna(0) <= 0)
                | (df_normal[HEIGHT_CM_COLUMN].fillna(0) <= 0)
            )
            if weight_missing_mask.any():
                for sku in df_normal.loc[weight_missing_mask, "SKU_ID"].unique():
                    missing_weight_skus.add(sku)

        df_normal["Mapped Name"] = df_normal[PRODUCT_NAME_COLUMN].fillna(df_normal["SKU_ID"])
        df_normal["Product Category"] = df_normal[PRODUCT_CATEGORY_COLUMN].fillna(df_normal["Mapped Name"])
        if is_local_store:
            df_normal["正常订单IVA_行"] = 0.0
        else:
            df_normal["正常订单IVA_行"] = df_normal.apply(
                lambda row: calc_mexico_iva_amount(
                    row.get(IMPORT_IVA_COLUMN, 0),
                    row[n_col] + row[p_col] + row[r_col],
                    row[ITEM_QUANTITY_COLUMN],
                ),
                axis=1,
            )
        if "Order ID" in df_normal.columns:
            daily_order_records.extend(
                df_normal[["日期", "Order ID"]]
                .dropna()
                .drop_duplicates()
                .assign(文件名=file_name)
                .to_dict("records")
            )

        sample_mask = ~status_series.isin(cancel_keywords) & (
            df[quantity_col].isna() | quantity_series.eq("") | quantity_series.eq("nan")
        )
        df_sample = df[sample_mask].copy()
        if not df_sample.empty:
            if "Paid Time" not in df_sample.columns:
                print(f"⚠️ 文件 {file_name} 样品订单中未找到 Paid Time 列，将跳过样品统计")
                df_sample = pd.DataFrame()
            else:
                df_sample["日期"] = df_sample["Paid Time"].map(parse_order_date)
                df_sample = df_sample.dropna(subset=["日期"]).copy()

        if not df_sample.empty:
            df_sample["SKU_ID"] = clean_sku_str(df_sample[SKU_ID_COLUMN])
            if ITEM_QUANTITY_COLUMN in df_sample.columns:
                df_sample[ITEM_QUANTITY_COLUMN] = pd.to_numeric(df_sample[ITEM_QUANTITY_COLUMN], errors="coerce").fillna(1)
            else:
                df_sample[ITEM_QUANTITY_COLUMN] = 1
            df_sample.loc[df_sample[ITEM_QUANTITY_COLUMN] <= 0, ITEM_QUANTITY_COLUMN] = 1
            df_sample = drop_conflicting_config_columns(df_sample, config_merge_cols)
            df_sample = df_sample.merge(config_df[config_merge_cols], left_on="SKU_ID", right_on=SKU_ID_COLUMN, how="left")
            missing_s = df_sample[df_sample[PRODUCT_NAME_COLUMN].isna()]
            if not missing_s.empty:
                for sku in missing_s["SKU_ID"].unique():
                    unmatched_skus.add(sku)
            sample_cost_missing_mask = df_sample["产品成本(元)"].fillna(0) <= 0
            if sample_cost_missing_mask.any():
                for sku in df_sample.loc[sample_cost_missing_mask, "SKU_ID"].unique():
                    missing_cost_skus.add(sku)
            if not is_local_store:
                sample_weight_missing_mask = (
                    (df_sample[ACTUAL_WEIGHT_KG_COLUMN].fillna(0) <= 0)
                    | (df_sample[LENGTH_CM_COLUMN].fillna(0) <= 0)
                    | (df_sample[WIDTH_CM_COLUMN].fillna(0) <= 0)
                    | (df_sample[HEIGHT_CM_COLUMN].fillna(0) <= 0)
                )
                if sample_weight_missing_mask.any():
                    for sku in df_sample.loc[sample_weight_missing_mask, "SKU_ID"].unique():
                        missing_weight_skus.add(sku)
            df_sample["Mapped Name"] = df_sample[PRODUCT_NAME_COLUMN].fillna(df_sample["SKU_ID"])
            df_sample["Product Category"] = df_sample[PRODUCT_CATEGORY_COLUMN].fillna(df_sample["Mapped Name"])

        normal_agg = {
            n_col: "sum",
            p_col: "sum",
            p_display_col: "sum",
            r_col: "sum",
            "产品成本(元)": "first",
            ITEM_QUANTITY_COLUMN: "sum",
            "正常订单IVA_行": "sum",
        }
        if not is_local_store and WEIGHT_KG_COLUMN in df_normal.columns:
            normal_agg[WEIGHT_KG_COLUMN] = "first"
        normal_daily_product = df_normal.groupby(["日期", "Product Category", "Mapped Name"]).agg(normal_agg).reset_index()
        normal_daily_product.rename(columns={ITEM_QUANTITY_COLUMN: "销量"}, inplace=True)
        if "Order ID" in df_normal.columns:
            sku_daily_order_count = (
                df_normal[["日期", "Product Category", "Mapped Name", "Order ID"]]
                .dropna(subset=["Order ID"])
                .assign(Order_ID_Clean=lambda x: x["Order ID"].astype(str).str.strip())
            )
            sku_daily_order_count = (
                sku_daily_order_count[sku_daily_order_count["Order_ID_Clean"] != ""]
                .groupby(["日期", "Product Category", "Mapped Name"])["Order_ID_Clean"]
                .nunique()
                .reset_index()
                .rename(columns={"Order_ID_Clean": "订单数"})
            )
        else:
            sku_daily_order_count = normal_daily_product[["日期", "Product Category", "Mapped Name", "销量"]].copy()
            sku_daily_order_count.rename(columns={"销量": "订单数"}, inplace=True)

        if is_local_store:
            normal_daily_product["销售额"] = normal_daily_product[n_col] + normal_daily_product[p_col]
        else:
            normal_daily_product["销售额"] = normal_daily_product[n_col] + normal_daily_product[p_col] + normal_daily_product[r_col]
        normal_daily_product["产品成本"] = normal_daily_product["销量"] * normal_daily_product["产品成本(元)"]
        if is_local_store:
            df_normal_local = build_mx_local_unit_logistics(df_normal)
            logistics_daily_product = build_order_logistics_allocation(df_normal_local, LOCAL_LOGISTICS_UNIT_COLUMN)
        else:
            logistics_daily_product = build_mexico_direct_logistics_allocation(df_normal, exchange_rate)
        normal_daily_product = normal_daily_product.merge(
            logistics_daily_product,
            on=["日期", "Product Category", "Mapped Name"],
            how="left",
        )
        normal_daily_product["物流成本"] = normal_daily_product["物流成本"].fillna(0)

        sample_daily_product = pd.DataFrame(columns=["日期", "Product Category", "Mapped Name", "寄样数", "寄样销售额", "寄样支出", "寄样IVA", "寄样总成本", "寄样除运费外销售额"])
        if not df_sample.empty:
            for col in [n_col, p_col, r_col]:
                df_sample[col] = df_sample[col].map(parse_amount)
            df_sample["寄样除运费外销售额"] = df_sample[n_col] + df_sample[p_col]
            sample_logistics_source_col = None
            if is_local_store:
                sample_logistics_source_col = LOGISTICS_COST_COLUMN if LOGISTICS_COST_COLUMN in df_sample.columns else HEAD_LOGISTICS_COST_COLUMN if HEAD_LOGISTICS_COST_COLUMN in df_sample.columns else None
            sample_agg = {
                n_col: "sum",
                p_col: "sum",
                "寄样除运费外销售额": "sum",
                r_col: "sum",
                "产品成本(元)": "first",
                ITEM_QUANTITY_COLUMN: "sum",
            }
            if not is_local_store and WEIGHT_KG_COLUMN in df_sample.columns:
                sample_agg[WEIGHT_KG_COLUMN] = "first"
            if sample_logistics_source_col and sample_logistics_source_col in df_sample.columns:
                sample_agg[sample_logistics_source_col] = "first"
            if is_local_store and HEAD_LOGISTICS_COST_COLUMN in df_sample.columns and TAIL_LOGISTICS_COST_COLUMN in df_sample.columns:
                sample_agg[HEAD_LOGISTICS_COST_COLUMN] = "first"
                sample_agg[TAIL_LOGISTICS_COST_COLUMN] = "first"
            if is_local_store:
                df_sample["寄样IVA_行"] = 0.0
            else:
                df_sample["寄样IVA_行"] = df_sample.apply(
                    lambda row: calc_mexico_iva_amount(
                        row.get(IMPORT_IVA_COLUMN, 0),
                        row[p_col],
                        row[ITEM_QUANTITY_COLUMN],
                    ),
                    axis=1,
                )
                sample_agg["寄样IVA_行"] = "sum"
            if is_local_store:
                sample_agg["寄样IVA_行"] = "sum"
            sample_daily_product = df_sample.groupby(["日期", "Product Category", "Mapped Name"]).agg(sample_agg).reset_index()
            sample_daily_product.rename(columns={ITEM_QUANTITY_COLUMN: "寄样数"}, inplace=True)
            if is_local_store:
                sample_daily_product["寄样销售额"] = sample_daily_product[n_col] + sample_daily_product[p_col]
            else:
                sample_daily_product["寄样销售额"] = sample_daily_product[n_col] + sample_daily_product[p_col] + sample_daily_product[r_col]
            if is_local_store:
                head_series = pd.to_numeric(sample_daily_product[HEAD_LOGISTICS_COST_COLUMN], errors="coerce").fillna(0) if HEAD_LOGISTICS_COST_COLUMN in sample_daily_product.columns else pd.Series(0, index=sample_daily_product.index)
                tail_series = pd.to_numeric(sample_daily_product[TAIL_LOGISTICS_COST_COLUMN], errors="coerce").fillna(0) if TAIL_LOGISTICS_COST_COLUMN in sample_daily_product.columns else pd.Series(0, index=sample_daily_product.index)
                sample_unit_cost = pd.to_numeric(sample_daily_product["产品成本(元)"], errors="coerce").fillna(0) + head_series + tail_series
                sample_daily_product["寄样支出"] = sample_daily_product["寄样数"] * sample_unit_cost
                sample_daily_product["寄样IVA"] = 0.0
                sample_daily_product["寄样总成本"] = sample_daily_product["寄样支出"]
            else:
                sample_logistics = build_mexico_direct_logistics_allocation(df_sample, exchange_rate)
                sample_daily_product = sample_daily_product.merge(
                    sample_logistics.rename(columns={"物流成本": "寄样物流成本"}),
                    on=["日期", "Product Category", "Mapped Name"],
                    how="left",
                )
                sample_daily_product["寄样物流成本"] = sample_daily_product["寄样物流成本"].fillna(0)
                sample_daily_product["寄样支出"] = (
                    sample_daily_product["寄样数"] * sample_daily_product["产品成本(元)"]
                    + sample_daily_product["寄样物流成本"]
                )
                sample_daily_product["寄样IVA"] = sample_daily_product["寄样IVA_行"].fillna(0) * exchange_rate
                sample_daily_product["寄样总成本"] = sample_daily_product["寄样支出"] + sample_daily_product["寄样IVA"]
            if is_local_store:
                sample_daily_product["寄样IVA"] = sample_daily_product["寄样IVA_行"].fillna(0)
            sample_daily_product = sample_daily_product[["日期", "Product Category", "Mapped Name", "寄样数", "寄样销售额", "寄样支出", "寄样IVA", "寄样总成本", "寄样除运费外销售额"]]

        merged_daily_product = normal_daily_product[[
            "日期", "Product Category", "Mapped Name", "销量", "销售额", p_col, p_display_col, "产品成本", "物流成本", "正常订单IVA_行"
        ]].merge(
            sample_daily_product,
            on=["日期", "Product Category", "Mapped Name"],
            how="outer",
        )
        if "寄样销售额" in merged_daily_product.columns:
            merged_daily_product["销售额"] = merged_daily_product["销售额"].fillna(0) + merged_daily_product["寄样销售额"].fillna(0)
        if "寄样除运费外销售额" in merged_daily_product.columns:
            merged_daily_product[p_display_col] = merged_daily_product[p_display_col].fillna(0) + merged_daily_product["寄样除运费外销售额"].fillna(0)
        if "正常订单IVA_行" in merged_daily_product.columns:
            merged_daily_product = merged_daily_product.rename(columns={"正常订单IVA_行": "正常订单IVA"})
        for col in ["销量", "销售额", p_col, p_display_col, "产品成本", "物流成本", "正常订单IVA", "寄样数", "寄样支出", "寄样IVA", "寄样总成本"]:
            if col in merged_daily_product.columns:
                merged_daily_product[col] = pd.to_numeric(merged_daily_product[col], errors="coerce").fillna(0)
        if is_local_store:
            merged_daily_product["正常订单IVA"] = 0
            merged_daily_product["寄样IVA"] = 0
            merged_daily_product["寄样总成本"] = merged_daily_product["寄样支出"]
        else:
            merged_daily_product["正常订单IVA"] = merged_daily_product["正常订单IVA"] * exchange_rate
            merged_daily_product["寄样总成本"] = merged_daily_product["寄样支出"] + merged_daily_product["寄样IVA"]
        merged_daily_product["汇率后金额"] = (merged_daily_product["销售额"] * exchange_rate).round(2)

        for _, row in merged_daily_product.iterrows():
            daily_product_detail_list.append({
                "文件名": file_name,
                "日期": row["日期"],
                "产品大类": row["Product Category"],
                "产品名称": row["Mapped Name"],
                "销量": int(row["销量"]),
                "销售额": round(row["销售额"], 2),
                "汇率后金额": round(row["汇率后金额"], 2),
                "除运费外销售额": round(row[p_display_col], 2),
                "产品成本": round(row["产品成本"], 2),
                "物流成本": round(row["物流成本"], 2),
                "正常订单IVA": round(row["正常订单IVA"], 2),
                "寄样数": int(row["寄样数"]),
                "寄样支出": round(row["寄样支出"], 2),
                "寄样IVA": round(row["寄样IVA"], 2),
                "寄样总成本": round(row["寄样总成本"], 2),
            })

        for _, row in sku_daily_order_count.iterrows():
            sku_order_detail_list.append({
                "文件名": file_name,
                "日期": row["日期"],
                "产品大类": row["Product Category"],
                "产品名称": row["Mapped Name"],
                "订单数": int(row["订单数"]),
            })

        for _, row in normal_daily_product.iterrows():
            product_quantity_records.append({
                "文件名": file_name,
                "日期": row["日期"],
                "产品名称": row["Product Category"],
                "销量": int(row["销量"]),
            })

        for _, row in sample_daily_product.iterrows():
            sample_quantity_records.append({
                "文件名": file_name,
                "日期": row["日期"],
                "产品名称": row["Product Category"],
                "寄样数": int(row["寄样数"]),
            })

    if not daily_product_detail_list:
        print(f"⚠️ [{display_name}] 没有有效数据，不生成报表。")
        return

    df_sku_detail = pd.DataFrame(daily_product_detail_list)
    if sku_order_detail_list:
        df_sku_order_count = pd.DataFrame(sku_order_detail_list).groupby(["文件名", "日期", "产品大类", "产品名称"])["订单数"].sum().reset_index()
        df_sku_detail = df_sku_detail.merge(df_sku_order_count, on=["文件名", "日期", "产品大类", "产品名称"], how="left")
    else:
        df_sku_detail["订单数"] = 0
    df_sku_detail["订单数"] = df_sku_detail["订单数"].fillna(0).astype(int)

    df_daily_product = df_sku_detail.groupby(["文件名", "日期", "产品大类"]).agg({
        "销量": "sum",
        "销售额": "sum",
        "汇率后金额": "sum",
        "除运费外销售额": "sum",
        "产品成本": "sum",
        "物流成本": "sum",
        "正常订单IVA": "sum",
        "寄样数": "sum",
        "寄样支出": "sum",
        "寄样IVA": "sum",
        "寄样总成本": "sum",
    }).reset_index().rename(columns={"产品大类": "产品名称"})

    df_daily_summary = df_daily_product.groupby(["文件名", "日期"]).agg({
        "销量": "sum",
        "销售额": "sum",
        "汇率后金额": "sum",
        "除运费外销售额": "sum",
        "产品成本": "sum",
        "物流成本": "sum",
        "正常订单IVA": "sum",
        "寄样数": "sum",
        "寄样支出": "sum",
        "寄样IVA": "sum",
        "寄样总成本": "sum",
    }).reset_index()
    if daily_order_records:
        df_order_count = pd.DataFrame(daily_order_records).drop_duplicates(subset=["文件名", "日期", "Order ID"]).groupby(["文件名", "日期"])["Order ID"].count().reset_index()
        df_order_count.rename(columns={"Order ID": "订单数"}, inplace=True)
    else:
        df_order_count = df_daily_product[df_daily_product["销量"] > 0].groupby(["文件名", "日期"])["销量"].sum().reset_index()
        df_order_count.rename(columns={"销量": "订单数"}, inplace=True)
    df_daily_summary = df_daily_summary.merge(df_order_count, on=["文件名", "日期"], how="left")
    df_daily_summary["订单数"] = df_daily_summary["订单数"].fillna(0).astype(int)
    df_daily_summary["每件商品成交费用"] = df_daily_summary["销量"] * MX_PER_ORDER_ITEM_FEE_MXN * exchange_rate
    if is_local_store:
        df_daily_summary["利润"] = (
            df_daily_summary["汇率后金额"]
            - df_daily_summary["产品成本"]
            - df_daily_summary["物流成本"]
            - df_daily_summary["寄样支出"]
            - df_daily_summary["每件商品成交费用"]
        )
    else:
        df_daily_summary["利润"] = (
            df_daily_summary["汇率后金额"]
            - df_daily_summary["产品成本"]
            - df_daily_summary["物流成本"]
            - df_daily_summary["正常订单IVA"]
            - df_daily_summary["寄样总成本"]
            - df_daily_summary["每件商品成交费用"]
        )
    df_daily_summary = df_daily_summary[[
        "文件名", "日期", "订单数", "销售额", "汇率后金额", "除运费外销售额", "产品成本", "物流成本", "正常订单IVA", "寄样支出", "寄样IVA", "寄样总成本", "每件商品成交费用", "利润", "销量", "寄样数"
    ]]
    df_daily_summary = df_daily_summary.assign(temp_date=pd.to_datetime(df_daily_summary["日期"], errors="coerce")) \
        .sort_values(["文件名", "temp_date"]) \
        .drop(columns=["temp_date"])
    total_row = {"文件名": "全部文件合计", "日期": "汇总"}
    for col in df_daily_summary.columns:
        if col not in ["文件名", "日期"]:
            total_row[col] = df_daily_summary[col].sum()
    df_daily_summary = pd.concat([df_daily_summary, pd.DataFrame([total_row])], ignore_index=True)

    df_daily_product = df_daily_product.assign(temp_date=pd.to_datetime(df_daily_product["日期"], errors="coerce")) \
        .sort_values(["temp_date", "文件名", "产品名称"]) \
        .drop(columns=["temp_date"])
    df_daily_product = df_daily_product[[
        "文件名", "日期", "产品名称", "销售额", "汇率后金额", "除运费外销售额", "产品成本", "物流成本", "正常订单IVA", "寄样支出", "寄样IVA", "寄样总成本", "销量", "寄样数"
    ]]
    df_sku_detail = df_sku_detail.assign(temp_date=pd.to_datetime(df_sku_detail["日期"], errors="coerce")) \
        .sort_values(["temp_date", "文件名", "产品大类", "产品名称"]) \
        .drop(columns=["temp_date"])
    df_sku_detail = df_sku_detail[[
        "文件名", "日期", "产品大类", "产品名称", "订单数", "销售额", "汇率后金额", "除运费外销售额", "产品成本", "物流成本", "正常订单IVA", "寄样支出", "寄样IVA", "寄样总成本", "销量", "寄样数"
    ]]

    period_comparison_frames = build_period_comparison_frames(df_daily_product, daily_order_records, exchange_rate)
    df_product_profit_by_period = build_product_profit_by_period(df_daily_product, exchange_rate)
    df_product_quantity_by_period = build_product_quantity_by_period(product_quantity_records)
    df_daily_product_quantity_matrix = build_daily_product_quantity_matrix(df_daily_product)
    df_daily_sku_quantity_matrix = build_daily_sku_quantity_matrix(df_sku_detail)
    df_sku_profit_by_period = build_sku_profit_by_period(df_sku_detail, exchange_rate)

    df_quantity_records = pd.DataFrame(product_quantity_records)
    if not df_quantity_records.empty:
        df_quantity_matrix = df_quantity_records.pivot_table(
            index=["文件名", "产品名称"],
            columns="日期",
            values="销量",
            aggfunc="sum",
            fill_value=0,
        ).astype(int)
        sorted_cols = sorted(df_quantity_matrix.columns, key=lambda x: pd.to_datetime(x, errors="coerce"))
        df_quantity_matrix = df_quantity_matrix[sorted_cols]
        df_quantity_matrix["汇总"] = df_quantity_matrix.sum(axis=1)
        total_row = df_quantity_matrix.sum(axis=0).to_frame().T
        total_row.index = pd.MultiIndex.from_tuples([("全部文件合计", "汇总")], names=df_quantity_matrix.index.names)
        df_quantity_matrix = pd.concat([df_quantity_matrix, total_row]).reset_index()
    else:
        df_quantity_matrix = pd.DataFrame([["无销量数据"]], columns=["提示"])

    df_sample_records = pd.DataFrame(sample_quantity_records)
    if not df_sample_records.empty:
        df_sample_summary = df_sample_records.groupby(["文件名", "产品名称"])["寄样数"].sum().reset_index()
        df_sample_summary = df_sample_summary.rename(columns={"寄样数": "汇总"})
        df_sample_matrix = df_sample_records.pivot_table(
            index=["文件名", "产品名称"],
            columns="日期",
            values="寄样数",
            aggfunc="sum",
            fill_value=0,
        ).astype(int)
        sorted_cols = sorted(df_sample_matrix.columns, key=lambda x: pd.to_datetime(x, errors="coerce"))
        df_sample_matrix = df_sample_matrix[sorted_cols].reset_index()
    else:
        df_sample_summary = pd.DataFrame([["无样品数据"]], columns=["提示"])
        df_sample_matrix = pd.DataFrame([["无样品数据"]], columns=["提示"])

    try:
        os.makedirs(os.path.dirname(output_filename), exist_ok=True)
        with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
            display_rename_map = {"寄样支出": "样品支出"}
            display_drop_cols = ["正常订单IVA", "寄样IVA", "除运费外销售额"] if is_local_store else []
            file_summary_sheet = "文件汇总"
            df_file_summary_display = df_product_profit_by_period.rename(columns=display_rename_map)
            if display_drop_cols:
                df_file_summary_display = df_file_summary_display.drop(columns=display_drop_cols, errors="ignore")
            df_file_summary_display = insert_blank_rows_between_files(df_file_summary_display)
            df_file_summary_display.to_excel(writer, sheet_name=file_summary_sheet, index=False)
            center_excel_sheet(writer, file_summary_sheet, len(df_file_summary_display) + 1, len(df_file_summary_display.columns))

            detail_sheet = "日产品明细"
            df_daily_product_quantity_matrix.to_excel(writer, sheet_name=detail_sheet, index=False)
            center_excel_sheet(writer, detail_sheet, len(df_daily_product_quantity_matrix) + 1, len(df_daily_product_quantity_matrix.columns))

            sku_detail_sheet = "SKU明细"
            df_daily_sku_quantity_matrix.to_excel(writer, sheet_name=sku_detail_sheet, index=False)
            center_excel_sheet(writer, sku_detail_sheet, len(df_daily_sku_quantity_matrix) + 1, len(df_daily_sku_quantity_matrix.columns))

            quantity_sheet = "产品销量矩阵"
            df_product_quantity_by_period.to_excel(writer, sheet_name=quantity_sheet, index=True)
            quantity_startrow = len(df_product_quantity_by_period) + 3
            df_quantity_matrix_display = insert_blank_rows_between_files(df_quantity_matrix)
            df_quantity_matrix_display.to_excel(writer, sheet_name=quantity_sheet, startrow=quantity_startrow, index=False)
            quantity_rows = quantity_startrow + len(df_quantity_matrix_display) + 1
            quantity_cols = max(len(df_product_quantity_by_period.columns) + 1, len(df_quantity_matrix_display.columns))
            center_excel_sheet(writer, quantity_sheet, quantity_rows, quantity_cols)

            sample_sheet = "样品统计"
            if "df_sample_summary" in locals() and "提示" not in df_sample_summary.columns:
                df_sample_summary_display = insert_blank_rows_between_files(df_sample_summary)
                df_sample_summary_display.to_excel(writer, sheet_name=sample_sheet, index=False)
                startrow = len(df_sample_summary_display) + 3
                df_sample_matrix.to_excel(writer, sheet_name=sample_sheet, startrow=startrow, index=False)
                total_rows = startrow + len(df_sample_matrix) + 1
                total_cols = max(len(df_sample_summary_display.columns), len(df_sample_matrix.columns))
                center_excel_sheet(writer, sample_sheet, total_rows, total_cols)
            else:
                df_sample_matrix.to_excel(writer, sheet_name=sample_sheet, index=False)
                center_excel_sheet(writer, sample_sheet, len(df_sample_matrix) + 1, len(df_sample_matrix.columns))

            if period_comparison_frames:
                startrow = 0
                sheet_name = "周期对比"
                max_cols = 0
                for title, frame in period_comparison_frames:
                    frame_display = frame.rename(columns=display_rename_map)
                    if display_drop_cols:
                        frame_display = frame_display.drop(columns=display_drop_cols, errors="ignore")
                    frame_display.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
                    max_cols = max(max_cols, len(frame_display.columns))
                    startrow += len(frame_display) + 3
                center_excel_sheet(writer, sheet_name, startrow, max_cols)
            else:
                period_hint = pd.DataFrame([{"提示": "周期对比需要至少 2 个有效 CSV 文件"}])
                period_hint.to_excel(writer, sheet_name="周期对比", index=False)
                center_excel_sheet(writer, "周期对比", len(period_hint) + 1, len(period_hint.columns))

        print(f"✅ [{display_name}] 处理完成，报表已保存至 {output_filename}")
        if unmatched_skus:
            print(f"💡 [{display_name}] 提醒：以下 SKU ID 在配置表中未找到匹配项:")
            for sku in sorted(list(unmatched_skus)):
                print(f"   - {sku}")
        if missing_cost_skus:
            print(f"💡 [{display_name}] 提醒：以下 SKU 在配置表中缺少产品成本，已按 0 处理:")
            for sku in sorted(list(missing_cost_skus)):
                print(f"   - {sku}")
        if missing_weight_skus:
            print(f"💡 [{display_name}] 提醒：以下 SKU 缺少实重/长宽高信息，已按 0 运费处理，请补飞书配置:")
            for sku in sorted(list(missing_weight_skus)):
                print(f"   - {sku}")
    except Exception as e:
        print(f"❌ [{display_name}] 保存失败: {e}")


def run_mexico_daily(stores=None):
    print("=" * 60)
    print("TikTok 墨西哥直邮按日汇总订单成本报表工具")
    print("=" * 60)

    exchange_rate = get_mx_exchange_rate()
    print(f"✅ 使用固定墨西哥汇率: {exchange_rate}")

    for store in (stores or MEXICO_STORES):
        if not store.get("enabled", True):
            continue

        sheet_name = store["sheet_name"]
        display_name = f"{store['country_name']}_{store['store_name']}"
        print(f"\n📋 处理组合: {display_name} (汇率: {exchange_rate})")

        try:
            config_df = get_config_dataframe(sheet_name)
            print(f"   ✅ 配置表加载成功，共 {len(config_df)} 条记录")
            run_report(store, config_df, exchange_rate)
        except Exception as e:
            print(f"   ❌ 处理失败: {e}")
            if "没有 SKU 数据" in str(e) or "无法获取配置数据" in str(e):
                print_order_sku_preview(store)
            continue

    print("\n" + "=" * 60)
    print("✨ 墨西哥站所有店铺处理完毕。")
    print("=" * 60)


if __name__ == "__main__":
    run_mexico_daily()
